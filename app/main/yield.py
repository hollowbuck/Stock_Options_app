#!/usr/bin/env python3
"""
OPTIMIZED ZERODHA OPTIONS PROCESSOR - Enhanced Version with FIXED ISSUES
Keeps all core logic from original yield.py while adding performance optimizations
Author: AI Assistant based on original yield.py
Version: 2.3 - Optimized Performance and Efficiency

Key optimizations applied:
1. Reduced unnecessary imports and variables
2. Optimized API call patterns and caching
3. Streamlined processing logic
4. Improved memory management
5. Enhanced output directory handling
6. Removed unused code and redundant operations
"""

import pandas as pd
import requests
import asyncio
try:
    import aiohttp
    AIOHTTP_AVAILABLE = True
except Exception:
    AIOHTTP_AVAILABLE = False
import threading
import time
import os
import random
import gc
from datetime import datetime
from io import StringIO
from concurrent.futures import ThreadPoolExecutor, as_completed
import difflib
import sqlite3

# Additional imports for formula preservation
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# Optional imports for real-time functionality
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False

try:
    import pythoncom  # For COM init on Windows threads
    PYTHONCOM_AVAILABLE = True
except Exception:
    PYTHONCOM_AVAILABLE = False

# Thread safety for printing - DEFINED FIRST
print_lock = threading.Lock()
data_collection_lock = threading.Lock()

def safe_print(*args, **kwargs):
    """Thread-safe printing with timestamps - MUST BE DEFINED BEFORE USAGE"""
    with print_lock:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{now}]", *args, **kwargs)

# Load credentials from environment variables (secure)
def read_zerodha_config():
    """
    Read Zerodha credentials from environment variables
    
    Required environment variables:
    - ZERODHA_API_KEY
    - ZERODHA_ACCESS_TOKEN
    
    Returns:
        dict with 'api_key' and 'access_token'
    
    Raises:
        ValueError if credentials are not configured
    """
    api_key = os.environ.get('ZERODHA_API_KEY')
    access_token = os.environ.get('ZERODHA_ACCESS_TOKEN')
    
    if not api_key or not access_token:
        raise ValueError(
            "Zerodha credentials not configured in environment variables. "
            "Please set ZERODHA_API_KEY and ZERODHA_ACCESS_TOKEN environment variables."
        )
    
    return {
        'api_key': api_key,
        'access_token': access_token
    }

# Load credentials at module level
try:
    credentials = read_zerodha_config()
    API_KEY = credentials['api_key']
    ACCESS_TOKEN = credentials['access_token']
except ValueError as e:
    # Don't raise immediately - allow app to start, but credentials won't work
    print(f"⚠️  Warning: {e}")
    API_KEY = None
    ACCESS_TOKEN = None
except Exception as e:
    print(f"❌ Error loading Zerodha credentials: {e}")
    API_KEY = None
    ACCESS_TOKEN = None

KITE_BASE_URL = "https://api.kite.trade"

# Configuration
MAIN_LIST_PATH = "Main_List.xlsx"
MAIN_LIST_DB = "Main_List.db"
OUTPUT_DIR = "Processed_Options"
OUTPUT_EXCEL_FILE = os.path.join(OUTPUT_DIR, "Options_Data.xlsx")
OUTPUT_DB_FILE = os.path.join(OUTPUT_DIR, "Options_Data.db")

# OPTIMIZED Performance settings - Streamlined
MAX_SYMBOL_WORKERS = 12  # Optimized for better throughput
MAX_OPTION_WORKERS_PER_SYMBOL = 6  # Balanced for stability
API_CALLS_PER_SECOND = 10  # Optimized rate limit
MAX_RETRIES_PER_REQUEST = 2  # Reduced for speed
BASE_RETRY_DELAY = 0.3  # Faster retry
MAX_RETRY_DELAY = 5.0  # Reduced max delay
BACKOFF_FACTOR = 1.3  # Faster backoff
JITTER_ENABLED = True

# Real-time settings
REAL_TIME_UPDATE_INTERVAL = 45  # Slightly faster updates
WEBSOCKET_ENABLED = True
USE_ASYNC_FETCH = False  # Default off to ensure stability; can be enabled later

# Circuit breaker settings - Streamlined
CIRCUIT_BREAKER_FAILURE_THRESHOLD = 8  # Reduced threshold
RECOVERY_PAUSE_TIME = 20  # Faster recovery
MAX_RECOVERY_CYCLES = 2  # Reduced cycles
CONCURRENCY_REDUCTION_FACTOR = 0.7  # More aggressive reduction

# Strike selection settings
STRIKE_PERCENTAGE_RANGE = 0.10  # ±10% range

# Invalid symbols to skip
INVALID_OPTION_SYMBOLS = {
    'NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY', 'NIFTYNXT50',
    'BSX'
}

# Define our script's data columns (these will be updated, others preserved)
SCRIPT_DATA_COLUMNS = [
    "Symbol", "CMP", "ScripName", "OptionType", "StrikePrice",
    "BidPrice", "BidQty", "AskPrice", "AskQty", "LTP",
    "Lot_Size", "Margin_Required", "LastUpdate", "Days"
]

# Global variables
selected_expiry_global = None
_instruments_cache = None
_current_data_cache = {}
_initial_processing_complete = False  # Flag to track initial processing

class AdvancedSymbolMatcher:
    """Optimized symbol matcher with essential functionality"""
    
    def __init__(self, instruments_cache):
        self.instruments_cache = instruments_cache
        self.nse_equity_symbols = set()
        self.symbol_variations = {}
        self.special_mappings = {
            'M&M': ['MM', 'MAHINDRA'],
            'L&T': ['LT', 'LARSENTOUBRO'],
            'H&R JOHNSON': ['HRJOHNSON'],
            'D-LINK': ['DLINK'],
            'V-MART': ['VMART']
        }
        
        if instruments_cache:
            self._build_database()
    
    def _build_database(self):
        """Build optimized symbol database"""
        if 'nse_equity' in self.instruments_cache:
            nse_df = self.instruments_cache['nse_equity']
            self.nse_equity_symbols = set(nse_df['tradingsymbol'].str.upper())
            
            # Build essential variations only
            for _, row in nse_df.iterrows():
                symbol = row['tradingsymbol'].upper()
                self.symbol_variations[symbol] = symbol
                
                # Add special character mappings
                for original, variations in self.special_mappings.items():
                    if original in self.nse_equity_symbols:
                        for variation in variations:
                            self.symbol_variations[variation] = original
                    for variation in variations:
                        if variation in self.nse_equity_symbols:
                            self.symbol_variations[original] = variation
                            break
            
            safe_print(f"✅ Symbol database built: {len(self.symbol_variations)} variations")
    
    def find_best_match(self, input_symbol):
        """Optimized symbol matching"""
        input_symbol = str(input_symbol).strip().upper()
        
        # Strategy 1: Exact match
        if input_symbol in self.nse_equity_symbols:
            return input_symbol
        
        # Strategy 2: Direct variation lookup
        if input_symbol in self.symbol_variations:
            return self.symbol_variations[input_symbol]
        
        # Strategy 3: Clean match
        input_clean = input_symbol.replace('&', '').replace(' ', '').replace('-', '')
        if input_clean in self.nse_equity_symbols:
            return input_clean
        
        # Strategy 4: Fuzzy matching
        close_matches = difflib.get_close_matches(
            input_symbol, list(self.nse_equity_symbols), n=2, cutoff=0.85
        )
        
        if close_matches:
            return close_matches[0]
        
        return None

class ThreadSafeRateLimiter:
    """Optimized rate limiter"""
    
    def __init__(self, calls_per_second=API_CALLS_PER_SECOND):
        self.base_calls_per_second = calls_per_second
        self.current_calls_per_second = calls_per_second
        self.last_call_time = 0
        self.timeout_count = 0
        self.lock = threading.Lock()
    
    def record_timeout(self):
        with self.lock:
            self.timeout_count += 1
            if self.timeout_count >= 3:  # Reduced threshold
                self.current_calls_per_second = max(1, self.current_calls_per_second * 0.8)
                self.timeout_count = 0
    
    def record_success(self):
        with self.lock:
            self.timeout_count = max(0, self.timeout_count - 1)
            if (self.current_calls_per_second < self.base_calls_per_second and 
                self.timeout_count == 0):
                self.current_calls_per_second = min(
                    self.base_calls_per_second, 
                    self.current_calls_per_second * 1.05  # Slower recovery
                )
    
    def wait_if_needed(self):
        with self.lock:
            if self.current_calls_per_second <= 0:
                return
            
            min_interval = 1.0 / self.current_calls_per_second
            current_time = time.time()
            
            if self.last_call_time > 0:
                time_since_last_call = current_time - self.last_call_time
                if time_since_last_call < min_interval:
                    sleep_time = min_interval - time_since_last_call
                    time.sleep(sleep_time)
            
            self.last_call_time = time.time()

class SmartCircuitBreaker:
    """Optimized circuit breaker"""
    
    def __init__(self):
        self.failure_count = 0
        self.timeout_count = 0
        self.recovery_cycles = 0
        self.is_recovering = False
        self.lock = threading.Lock()
    
    def record_failure(self):
        with self.lock:
            self.failure_count += 1
    
    def record_timeout(self):
        with self.lock:
            self.timeout_count += 1
    
    def record_success(self):
        with self.lock:
            self.failure_count = max(0, self.failure_count - 1)
            self.timeout_count = max(0, self.timeout_count - 1)
    
    def check_and_recover_if_needed(self, rate_limiter, current_workers):
        with self.lock:
            total_issues = self.failure_count + self.timeout_count
            
            if total_issues >= CIRCUIT_BREAKER_FAILURE_THRESHOLD and not self.is_recovering:
                return self._perform_recovery(rate_limiter, current_workers)
            
            return current_workers
    
    def _perform_recovery(self, rate_limiter, current_workers):
        self.is_recovering = True
        self.recovery_cycles += 1
        
        safe_print(f"🔄 RECOVERY #{self.recovery_cycles} - Pausing {RECOVERY_PAUSE_TIME}s...")
        time.sleep(RECOVERY_PAUSE_TIME)
        
        self.failure_count = 0
        self.timeout_count = 0
        self.is_recovering = False
        
        new_workers = max(1, int(current_workers * CONCURRENCY_REDUCTION_FACTOR))
        rate_limiter.current_calls_per_second = max(1, int(rate_limiter.base_calls_per_second * 0.6))
        
        safe_print(f"✅ Recovery complete: {current_workers} → {new_workers} workers")
        return new_workers

class OptimizedCMPCache:
    """Optimized CMP caching"""
    
    def __init__(self, ttl_seconds=300):  # 5 minutes TTL
        self.cache = {}
        self.timestamps = {}
        self.ttl = ttl_seconds
        self.lock = threading.Lock()
        self.max_size = 1500  # Reduced cache size
        self.hit_count = 0
        self.miss_count = 0
    
    def get(self, symbol):
        """Get CMP from cache if valid"""
        with self.lock:
            if symbol in self.cache:
                if time.time() - self.timestamps[symbol] < self.ttl:
                    self.hit_count += 1
                    return self.cache[symbol]
                else:
                    # Expired, remove
                    del self.cache[symbol]
                    del self.timestamps[symbol]
            
            self.miss_count += 1
            return None
    
    def set(self, symbol, cmp_value):
        """Set CMP in cache with cleanup"""
        with self.lock:
            # Cleanup if cache is too large
            if len(self.cache) >= self.max_size:
                # Remove oldest 25% of entries
                cleanup_count = int(self.max_size * 0.25)
                oldest_symbols = sorted(self.timestamps.items(), key=lambda x: x[1])[:cleanup_count]
                for old_symbol, _ in oldest_symbols:
                    self.cache.pop(old_symbol, None)
                    self.timestamps.pop(old_symbol, None)
            
            self.cache[symbol] = cmp_value
            self.timestamps[symbol] = time.time()
    
    def get_stats(self):
        """Get cache statistics"""
        total = self.hit_count + self.miss_count
        hit_rate = (self.hit_count / total * 100) if total > 0 else 0
        return {
            'size': len(self.cache),
            'hit_rate': hit_rate,
            'hits': self.hit_count,
            'misses': self.miss_count
        }
    
    def clear_expired_only(self):
        """Clear only expired entries"""
        current_time = time.time()
        with self.lock:
            expired_keys = [symbol for symbol, timestamp in self.timestamps.items() 
                          if current_time - timestamp >= self.ttl]
            
            for symbol in expired_keys:
                self.cache.pop(symbol, None)
                self.timestamps.pop(symbol, None)
    
    def clear(self):
        """Clear cache"""
        with self.lock:
            self.cache.clear()
            self.timestamps.clear()
            self.hit_count = 0
            self.miss_count = 0

class OptimizedExcelManager:
    """Optimized Excel operations"""
    
    def __init__(self):
        self.structure_cache = {}
        self.workbook_cache = None
        self.last_modified = 0
        self.lock = threading.Lock()
    
    def get_cached_structure(self, filename, sheet_name):
        """Get cached Excel structure"""
        cache_key = f"{filename}:{sheet_name}"
        file_mtime = os.path.getmtime(filename) if os.path.exists(filename) else 0
        
        with self.lock:
            if (cache_key in self.structure_cache and 
                self.structure_cache[cache_key]['mtime'] == file_mtime):
                return self.structure_cache[cache_key]['data']
        
        return None
    
    def cache_structure(self, filename, sheet_name, data):
        """Cache Excel structure"""
        cache_key = f"{filename}:{sheet_name}"
        file_mtime = os.path.getmtime(filename) if os.path.exists(filename) else 0
        
        with self.lock:
            self.structure_cache[cache_key] = {
                'data': data,
                'mtime': file_mtime
            }
            
            # Cleanup old cache entries
            if len(self.structure_cache) > 30:  # Reduced cache size
                oldest_keys = list(self.structure_cache.keys())[:5]
                for key in oldest_keys:
                    del self.structure_cache[key]
    
    def get_workbook(self, filename):
        """Get cached workbook or load new one"""
        file_mtime = os.path.getmtime(filename) if os.path.exists(filename) else 0
        
        with self.lock:
            if (self.workbook_cache and 
                self.last_modified == file_mtime):
                return self.workbook_cache
            
            # Load new workbook
            if os.path.exists(filename):
                self.workbook_cache = load_workbook(filename, data_only=False)
                self.last_modified = file_mtime
                return self.workbook_cache
        
        return None
    
    def clear_cache(self):
        """Clear all caches"""
        with self.lock:
            self.structure_cache.clear()
            if self.workbook_cache:
                self.workbook_cache.close()
            self.workbook_cache = None
            self.last_modified = 0

class OptimizedSessionPool:
    """Optimized connection pool"""
    
    def __init__(self, pool_size=8):  # Optimized pool size
        self.pool = []
        self.pool_size = pool_size
        self.lock = threading.Lock()
        self.created_count = 0
        self.reused_count = 0
        self._create_sessions()
    
    def _create_sessions(self):
        """Create initial session pool"""
        for _ in range(self.pool_size):
            session = requests.Session()
            session.headers.update({
                'X-Kite-Version': '3',
                'Content-Type': 'application/json',
                'Authorization': f'token {API_KEY}:{ACCESS_TOKEN}'
            })
            
            # Optimized session configuration
            adapter = requests.adapters.HTTPAdapter(
                pool_connections=10,
                pool_maxsize=20,
                max_retries=0
            )
            session.mount('http://', adapter)
            session.mount('https://', adapter)
            
            self.pool.append(session)
            self.created_count += 1
    
    def get_session(self):
        """Get session from pool"""
        with self.lock:
            if self.pool:
                self.reused_count += 1
                return self.pool.pop()
            else:
                # Create new session if pool is empty
                session = requests.Session()
                session.headers.update({
                    'X-Kite-Version': '3',
                    'Content-Type': 'application/json',
                    'Authorization': f'token {API_KEY}:{ACCESS_TOKEN}'
                })
                
                adapter = requests.adapters.HTTPAdapter(
                    pool_connections=5,
                    pool_maxsize=15,
                    max_retries=0
                )
                session.mount('http://', adapter)
                session.mount('https://', adapter)
                
                self.created_count += 1
                return session
    
    def return_session(self, session):
        """Return session to pool"""
        with self.lock:
            if len(self.pool) < self.pool_size:
                self.pool.append(session)
            else:
                session.close()
    
    def get_stats(self):
        """Get pool statistics"""
        return {
            'pool_size': len(self.pool),
            'max_size': self.pool_size,
            'created': self.created_count,
            'reused': self.reused_count
        }
    
    def close_all(self):
        """Close all sessions"""
        with self.lock:
            for session in self.pool:
                session.close()
            self.pool.clear()

# Global instances - Separate rate limiters for better performance
quote_rate_limiter = ThreadSafeRateLimiter(calls_per_second=API_CALLS_PER_SECOND)
margin_rate_limiter = ThreadSafeRateLimiter(calls_per_second=API_CALLS_PER_SECOND)
rate_limiter = quote_rate_limiter  # Default for backward compatibility
circuit_breaker = SmartCircuitBreaker()
symbol_matcher = None
cmp_cache = OptimizedCMPCache()
excel_manager = OptimizedExcelManager()
session_pool = OptimizedSessionPool()

# Async HTTP session pool
class AsyncSessionPool:
    def __init__(self, pool_size: int = 8):
        self.pool_size = pool_size
        self._sessions = []
        self._lock = threading.Lock()
        self._closed = False

    async def _create_session(self):
        timeout = aiohttp.ClientTimeout(total=15)
        headers = {
            'X-Kite-Version': '3',
            'Content-Type': 'application/json',
            'Authorization': f'token {API_KEY}:{ACCESS_TOKEN}'
        }
        return aiohttp.ClientSession(timeout=timeout, headers=headers)

    async def get(self) -> aiohttp.ClientSession:
        with self._lock:
            if self._sessions:
                return self._sessions.pop()
        # Create outside lock
        return await self._create_session()

    async def put(self, session: aiohttp.ClientSession):
        if self._closed:
            await session.close()
            return
        with self._lock:
            if len(self._sessions) < self.pool_size:
                self._sessions.append(session)
                return
        await session.close()

    async def close(self):
        self._closed = True
        with self._lock:
            sessions = list(self._sessions)
            self._sessions.clear()
        for s in sessions:
            try:
                await s.close()
            except Exception:
                pass

async_pool = None

# SQLite helpers
try:
    from .db_utils import (
        write_frames_as_tables,
        read_table,
        read_all_tables,
        upsert_symbols_list,
        read_symbols_list,
    )
except Exception:
    # Fallback for direct script execution
    from app.main.db_utils import (
        write_frames_as_tables,
        read_table,
        read_all_tables,
        upsert_symbols_list,
        read_symbols_list,
    )

def ensure_output_directory():
    """Ensure output directory exists"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        safe_print(f"📁 Created output directory: {OUTPUT_DIR}")

def calculate_days_to_expiry(expiry_date_str):
    """Calculate days remaining until expiry date"""
    try:
        if isinstance(expiry_date_str, str):
            if '-' in expiry_date_str:
                expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
            else:
                expiry_date = datetime.strptime(expiry_date_str, '%d-%m-%Y').date()
        else:
            expiry_date = expiry_date_str.date() if hasattr(expiry_date_str, 'date') else expiry_date_str
        
        current_date = datetime.now().date()
        days_remaining = (expiry_date - current_date).days
        
        # Ensure minimum of 1 day to avoid division by zero in calculations
        # This handles the case when today is the expiry day (returns 1 instead of 0)
        return max(1, days_remaining)
    except Exception:
        return 1

def create_zerodha_session():
    """Use session pool"""
    return session_pool.get_session()

def test_zerodha_authentication():
    """Test authentication"""
    try:
        session = create_zerodha_session()
        try:
            response = session.get(f'{KITE_BASE_URL}/user/profile', timeout=10)
            if response.status_code == 200:
                profile = response.json()['data']
                safe_print(f"✅ Authentication successful: {profile['user_name']}")
                return True
            else:
                safe_print(f"❌ Authentication failed: {response.text}")
                return False
        finally:
            session_pool.return_session(session)
    except Exception as e:
        safe_print(f"❌ Authentication error: {e}")
        return False

def fetch_instruments():
    """Optimized instruments fetch"""
    global _instruments_cache, symbol_matcher
    
    if _instruments_cache is not None:
        return _instruments_cache
    
    try:
        session = create_zerodha_session()
        try:
            rate_limiter.wait_if_needed()
            safe_print("📥 Fetching instruments from Zerodha API...")
            
            response = session.get(f'{KITE_BASE_URL}/instruments', timeout=30)
            
            if response.status_code == 200:
                df = pd.read_csv(StringIO(response.text))
                
                # Filter instruments
                nfo_options = df[
                    (df['exchange'] == 'NFO') &
                    (df['instrument_type'].isin(['CE', 'PE']))
                ].copy()
                
                nse_equity = df[
                    (df['exchange'] == 'NSE') &
                    (df['instrument_type'] == 'EQ')
                ].copy()
                
                _instruments_cache = {
                    'nfo_options': nfo_options,
                    'nse_equity': nse_equity,
                    'raw_data': df
                }
                
                # Initialize symbol matcher
                symbol_matcher = AdvancedSymbolMatcher(_instruments_cache)
                
                safe_print(f"✅ Instruments loaded: NFO={len(nfo_options)}, NSE={len(nse_equity)}")
                circuit_breaker.record_success()
                rate_limiter.record_success()
                return _instruments_cache
            else:
                safe_print(f"❌ Failed to fetch instruments: Status {response.status_code}")
                if response.status_code == 429:
                    rate_limiter.record_timeout()
                else:
                    circuit_breaker.record_failure()
                return None
        finally:
            session_pool.return_session(session)
    except requests.exceptions.Timeout:
        circuit_breaker.record_timeout()
        rate_limiter.record_timeout()
        safe_print("❌ Instruments fetch timeout")
        return None
    except Exception as e:
        circuit_breaker.record_failure()
        safe_print(f"❌ Error fetching instruments: {e}")
        return None

def validate_fo_availability(symbol):
    """Check if symbol has F&O contracts available"""
    instruments = fetch_instruments()
    if not instruments or 'nfo_options' not in instruments:
        return False
    
    nfo_options = instruments['nfo_options']
    
    # Check for exact symbol match
    fo_exists = len(nfo_options[
        nfo_options['name'].str.contains(f'^{symbol}$', case=False, na=False, regex=True)
    ]) > 0
    
    # Fallback: check word boundary match
    if not fo_exists:
        fo_exists = len(nfo_options[
            nfo_options['name'].str.contains(f'\\b{symbol}\\b', case=False, na=False, regex=True)
        ]) > 0
    
    return fo_exists

def fetch_cmp_zerodha(symbol):
    """Optimized CMP fetching with caching"""
    global symbol_matcher
    
    symbol_norm = symbol.strip().upper()
    
    # Check cache first
    cached_cmp = cmp_cache.get(symbol_norm)
    if cached_cmp is not None:
        return cached_cmp
    
    # Use symbol matcher
    if symbol_matcher:
        matched_symbol = symbol_matcher.find_best_match(symbol_norm)
        if not matched_symbol:
            return None
        symbol_norm = matched_symbol
    
    # Get session from pool
    session = session_pool.get_session()
    
    try:
        for attempt in range(MAX_RETRIES_PER_REQUEST):
            try:
                rate_limiter.wait_if_needed()
                response = session.get(
                    f'{KITE_BASE_URL}/quote', 
                    params={'i': f'NSE:{symbol_norm}'}, 
                    timeout=8
                )
                
                if response.status_code == 200:
                    data = response.json()['data']
                    if f'NSE:{symbol_norm}' in data:
                        ltp = data[f'NSE:{symbol_norm}']['last_price']
                        circuit_breaker.record_success()
                        rate_limiter.record_success()
                        
                        # Cache the result
                        cmp_cache.set(symbol_norm, ltp)
                        return ltp
                elif response.status_code == 429:
                    rate_limiter.record_timeout()
                else:
                    circuit_breaker.record_failure()
                if attempt < MAX_RETRIES_PER_REQUEST - 1:
                    delay = BASE_RETRY_DELAY * (BACKOFF_FACTOR ** attempt)
                    if JITTER_ENABLED:
                        delay *= (0.5 + random.random())
                    time.sleep(min(delay, MAX_RETRY_DELAY))
                    continue
                    
            except requests.exceptions.Timeout:
                circuit_breaker.record_timeout()
                rate_limiter.record_timeout()
                if attempt < MAX_RETRIES_PER_REQUEST - 1:
                    time.sleep(0.3 * (attempt + 1))
                    continue
            except Exception:
                circuit_breaker.record_failure()
                if attempt < MAX_RETRIES_PER_REQUEST - 1:
                    time.sleep(0.3 * (attempt + 1))
                    continue
            
            break
        
        return None
        
    finally:
        session_pool.return_session(session)

def get_global_expiry(symbol):
    """Get expiry selection"""
    global selected_expiry_global
    
    instruments = fetch_instruments()
    if not instruments or 'nfo_options' not in instruments:
        safe_print("❌ No NFO instruments found")
        return False
    
    nfo_options = instruments['nfo_options']
    expiry_dates = sorted(nfo_options['expiry'].dropna().unique())
    
    if len(expiry_dates) == 0:
        safe_print("❌ No expiries found")
        return False
    
    safe_print("\nAvailable expiries:")
    for i, date in enumerate(expiry_dates[:5]):
        safe_print(f"{i+1}. {date}")
    
    try:
        sel = input("Select expiry (or Enter for first): ").strip()
        if sel == "":
            selected_expiry_global = expiry_dates[0]
        else:
            idx = int(sel) - 1
            selected_expiry_global = expiry_dates[idx] if 0 <= idx < len(expiry_dates) else expiry_dates[0]
    except:
        selected_expiry_global = expiry_dates[0]
    
    safe_print(f"Selected expiry: {selected_expiry_global}")
    return True

def fetch_single_margin_with_validation(zerodha_symbol, lot_size):
    """Individual margin fetch with validation"""
    session = session_pool.get_session()
    
    try:
        for attempt in range(2):
            try:
                margin_rate_limiter.wait_if_needed()  # Use separate rate limiter for margins
                
                margin_order = [{
                    "exchange": "NFO",
                    "tradingsymbol": zerodha_symbol,
                    "transaction_type": "SELL",
                    "variety": "regular",
                    "product": "NRML",
                    "order_type": "MARKET",
                    "quantity": int(lot_size),
                    "price": 0,
                    "trigger_price": 0
                }]
                
                margin_response = session.post(
                    f'{KITE_BASE_URL}/margins/orders', 
                    json=margin_order,
                    timeout=8
                )
                
                if margin_response.status_code == 200:
                    margin_data = margin_response.json().get('data', [])
                    
                    if margin_data and len(margin_data) > 0:
                        first_item = margin_data[0]
                        if isinstance(first_item, dict):
                            total_margin = first_item.get('total', 0)
                            
                            if total_margin and total_margin > 10:
                                circuit_breaker.record_success()
                                rate_limiter.record_success()
                                return int(round(total_margin))
                                
                elif margin_response.status_code == 429:
                    rate_limiter.record_timeout()
                else:
                    circuit_breaker.record_failure()
                    
            except requests.exceptions.Timeout:
                rate_limiter.record_timeout()
            except Exception:
                circuit_breaker.record_failure()
            
            if attempt < 1:
                time.sleep(0.3)
        
        return 0
        
    finally:
        session_pool.return_session(session)

def fetch_option_data_optimized(zerodha_symbol, lot_size):
    """Optimized option data fetching"""
    session = session_pool.get_session()
    
    result = {
        'ltp': 0,
        'bid_price': 0,
        'bid_qty': 0,
        'ask_price': 0,
        'ask_qty': 0,
        'margin_required': 0
    }
    
    try:
        for attempt in range(2):
            try:
                rate_limiter.wait_if_needed()
                quote_response = session.get(
                    f'{KITE_BASE_URL}/quote', 
                    params={'i': f'NFO:{zerodha_symbol}'},
                    timeout=6
                )
                
                if quote_response.status_code == 200:
                    quote_data = quote_response.json()['data']
                    if f'NFO:{zerodha_symbol}' in quote_data:
                        option_data = quote_data[f'NFO:{zerodha_symbol}']
                        result['ltp'] = option_data.get('last_price', 0)
                        
                        depth = option_data.get('depth', {})
                        if 'buy' in depth and depth['buy']:
                            result['bid_price'] = depth['buy'][0].get('price', 0)
                            result['bid_qty'] = depth['buy'][0].get('quantity', 0)
                        
                        if 'sell' in depth and depth['sell']:
                            result['ask_price'] = depth['sell'][0].get('price', 0)
                            result['ask_qty'] = depth['sell'][0].get('quantity', 0)
                        
                        break
                elif quote_response.status_code == 429:
                    rate_limiter.record_timeout()
            except requests.exceptions.Timeout:
                rate_limiter.record_timeout()
            except Exception:
                pass
            
            if attempt < 1:
                time.sleep(0.2)
        
        circuit_breaker.record_success()
        rate_limiter.record_success()
        
    except Exception:
        circuit_breaker.record_failure()
    
    finally:
        session_pool.return_session(session)
    
    # Fetch margin if quote data is valid
    if result['ltp'] > 0 or result['bid_price'] > 0:
        result['margin_required'] = fetch_single_margin_with_validation(zerodha_symbol, lot_size)
    
    return result

async def fetch_quotes_batch_nfo_async(contract_symbols):
    if not contract_symbols or not AIOHTTP_AVAILABLE:
        return {}
    global async_pool
    if async_pool is None:
        async_pool = AsyncSessionPool(pool_size=8)

    session = await async_pool.get()
    try:
        result = {}
        chunk_size = 10
        for i in range(0, len(contract_symbols), chunk_size):
            chunk = contract_symbols[i:i+chunk_size]
            params = [("i", f"NFO:{sym}") for sym in chunk]
            for attempt in range(2):
                try:
                    quote_rate_limiter.wait_if_needed()
                    async with session.get(f"{KITE_BASE_URL}/quote", params=params) as resp:
                        if resp.status == 200:
                            data = (await resp.json()).get("data", {})
                            for sym in chunk:
                                key = f"NFO:{sym}"
                                if key in data:
                                    q = data[key]
                                    out = {
                                        'ltp': q.get('last_price', 0),
                                        'bid_price': 0,
                                        'bid_qty': 0,
                                        'ask_price': 0,
                                        'ask_qty': 0,
                                    }
                                    depth = q.get('depth', {}) or {}
                                    buys = depth.get('buy') or []
                                    sells = depth.get('sell') or []
                                    if buys:
                                        out['bid_price'] = buys[0].get('price', 0)
                                        out['bid_qty'] = buys[0].get('quantity', 0)
                                    if sells:
                                        out['ask_price'] = sells[0].get('price', 0)
                                        out['ask_qty'] = sells[0].get('quantity', 0)
                                    result[sym] = out
                            break
                        elif resp.status == 429:
                            quote_rate_limiter.record_timeout()
                        else:
                            circuit_breaker.record_failure()
                except asyncio.TimeoutError:
                    quote_rate_limiter.record_timeout()
                except Exception:
                    circuit_breaker.record_failure()
                await asyncio.sleep(0.25)
        if result:
            circuit_breaker.record_success()
            quote_rate_limiter.record_success()
        return result
    finally:
        await async_pool.put(session)

async def fetch_single_margin_async(zerodha_symbol: str, lot_size: int) -> int:
    if not AIOHTTP_AVAILABLE:
        return 0
    global async_pool
    if async_pool is None:
        async_pool = AsyncSessionPool(pool_size=8)
    session = await async_pool.get()
    try:
        for attempt in range(2):
            try:
                margin_rate_limiter.wait_if_needed()
                margin_order = [{
                    "exchange": "NFO",
                    "tradingsymbol": zerodha_symbol,
                    "transaction_type": "SELL",
                    "variety": "regular",
                    "product": "NRML",
                    "order_type": "MARKET",
                    "quantity": int(lot_size),
                    "price": 0,
                    "trigger_price": 0
                }]
                async with session.post(f'{KITE_BASE_URL}/margins/orders', json=margin_order) as resp:
                    if resp.status == 200:
                        margin_data = (await resp.json()).get('data', [])
                        if margin_data and isinstance(margin_data[0], dict):
                            total_margin = margin_data[0].get('total', 0)
                            if total_margin and total_margin > 10:
                                return int(round(total_margin))
                    elif resp.status == 429:
                        rate_limiter.record_timeout()
            except asyncio.TimeoutError:
                rate_limiter.record_timeout()
            except Exception:
                circuit_breaker.record_failure()
            await asyncio.sleep(0.3)
        return 0
    finally:
        await async_pool.put(session)

def fetch_margins_individually_parallel(contracts_info, max_workers=4):
    if not contracts_info:
        return {}
    results = {}
    def fetch_one(tsym, lot):
        return tsym, fetch_single_margin_with_validation(tsym, lot)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(fetch_one, tsym, lot): tsym for tsym, lot in contracts_info}
        for future in as_completed(futures):
            try:
                tsym, margin = future.result()
                results[tsym] = margin
            except Exception:
                continue
    return results

def fetch_quotes_batch_nfo(contract_symbols):
    """Optimized batch fetch quotes with adaptive batch sizing"""
    if not contract_symbols:
        return {}
    
    session = session_pool.get_session()
    
    try:
        result = {}
        
        # Adaptive chunk size - start larger, reduce if issues occur
        chunk_size = 15  # Increased from 10 to 15 for better throughput
        consecutive_timeouts = 0
        
        for i in range(0, len(contract_symbols), chunk_size):
            chunk = contract_symbols[i:i+chunk_size]
            params = [("i", f"NFO:{sym}") for sym in chunk]
            
            for attempt in range(2):
                try:
                    quote_rate_limiter.wait_if_needed()  # Use separate rate limiter for quotes
                    resp = session.get(f"{KITE_BASE_URL}/quote", params=params, timeout=10)
                    
                    if resp.status_code == 200:
                        data = resp.json().get("data", {})
                        for sym in chunk:
                            key = f"NFO:{sym}"
                            if key in data:
                                q = data[key]
                                out = {
                                    'ltp': q.get('last_price', 0),
                                    'bid_price': 0,
                                    'bid_qty': 0,
                                    'ask_price': 0,
                                    'ask_qty': 0,
                                }
                                
                                depth = q.get('depth', {}) or {}
                                buys = depth.get('buy') or []
                                sells = depth.get('sell') or []
                                
                                if buys:
                                    out['bid_price'] = buys[0].get('price', 0)
                                    out['bid_qty'] = buys[0].get('quantity', 0)
                                if sells:
                                    out['ask_price'] = sells[0].get('price', 0)
                                    out['ask_qty'] = sells[0].get('quantity', 0)
                                
                                result[sym] = out
                        break
                    elif resp.status_code == 429:
                        quote_rate_limiter.record_timeout()
                        # Do not count as failure for circuit breaker
                        consecutive_timeouts += 1
                        # Reduce chunk size if too many timeouts
                        if consecutive_timeouts >= 3:
                            chunk_size = max(5, chunk_size - 2)
                            consecutive_timeouts = 0
                    else:
                        circuit_breaker.record_failure()
                        
                except requests.exceptions.Timeout:
                    quote_rate_limiter.record_timeout()
                    consecutive_timeouts += 1
                    if consecutive_timeouts >= 3:
                        chunk_size = max(5, chunk_size - 2)
                        consecutive_timeouts = 0
                except Exception:
                    circuit_breaker.record_failure()
                
                if attempt < 1:
                    time.sleep(0.2)
            
            # Small pacing delay between chunks
            time.sleep(0.15)
            # Reset timeout counter on success
            consecutive_timeouts = 0
        
        if result:
            circuit_breaker.record_success()
            quote_rate_limiter.record_success()
        
        return result
    
    except Exception:
        circuit_breaker.record_failure()
        return {}
    finally:
        session_pool.return_session(session)

def fetch_option_chain_zerodha(symbol, cmp, current_option_workers):
    """Optimized option chain fetching"""
    global selected_expiry_global, symbol_matcher
    
    symbol_norm = symbol.strip().upper()
    
    # Use symbol matcher
    if symbol_matcher:
        matched_symbol = symbol_matcher.find_best_match(symbol_norm)
        if matched_symbol:
            symbol_norm = matched_symbol
    
    # Validate F&O availability
    if not validate_fo_availability(symbol_norm):
        return None
    
    instruments = fetch_instruments()
    if not instruments or 'nfo_options' not in instruments:
        return None
    
    nfo_options = instruments['nfo_options']
    
    # Filter options
    symbol_options = nfo_options[
        (nfo_options['name'].str.contains(f'\\b{symbol_norm}\\b', case=False, na=False, regex=True)) &
        (nfo_options['expiry'] == selected_expiry_global)
    ]
    
    # Fallback: exact match
    if len(symbol_options) == 0:
        symbol_options = nfo_options[
            (nfo_options['name'].str.contains(f'^{symbol_norm}$', case=False, na=False, regex=True)) &
            (nfo_options['expiry'] == selected_expiry_global)
        ]
    
    if len(symbol_options) == 0:
        return None
    
    # Apply strike selection logic
    delta = cmp * STRIKE_PERCENTAGE_RANGE
    lower_bound = cmp - delta
    upper_bound = cmp + delta
    
    selected_contracts = symbol_options[
        (
            (symbol_options['instrument_type'] == 'PE') &
            (symbol_options['strike'] < lower_bound)
        ) |
        (
            (symbol_options['instrument_type'] == 'CE') &
            (symbol_options['strike'] > upper_bound)
        )
    ]
    
    if len(selected_contracts) == 0:
        return None
    
    results = []
    contracts_list = selected_contracts.to_dict('records')
    
    # Try batch processing first (async preferred)
    symbols_list = [row['tradingsymbol'] for row in contracts_list]
    quotes = {}
    if USE_ASYNC_FETCH and AIOHTTP_AVAILABLE:
        try:
            quotes = asyncio.run(fetch_quotes_batch_nfo_async(symbols_list))
        except Exception:
            quotes = {}
    if not quotes:
        try:
            quotes = fetch_quotes_batch_nfo(symbols_list)
        except Exception:
            quotes = {}
    
    def process_contract_balanced(row):
        """Process contract with batch quotes + individual margin (parallel only - no filtering)."""
        try:
            tsym = row['tradingsymbol']
            lot_size = int(row['lot_size'])
            
            # Try to use batch quote data first
            q = quotes.get(tsym)
            
            # If batch data is missing, fetch individually
            if not q or (q.get('ltp', 0) == 0 and q.get('bid_price', 0) == 0):
                individual_data = fetch_option_data_optimized(tsym, lot_size)
                q = individual_data
                margin_val = individual_data.get('margin_required', 0)
            else:
                # Use batch quote data but fetch margin individually when contract has value
                margin_val = 0
                if q.get('ltp', 0) > 0 or q.get('bid_price', 0) > 0:
                    if USE_ASYNC_FETCH and AIOHTTP_AVAILABLE:
                        try:
                            margin_val = asyncio.run(fetch_single_margin_async(tsym, lot_size))
                        except Exception:
                            margin_val = fetch_single_margin_with_validation(tsym, lot_size)
                    else:
                        margin_val = fetch_single_margin_with_validation(tsym, lot_size)
            
            ltp = (q or {}).get('ltp', 0)
            bid = (q or {}).get('bid_price', 0)
            
            if ltp > 0 or bid > 0:
                days_to_expiry = calculate_days_to_expiry(selected_expiry_global)
                
                return {
                    "Symbol": symbol_norm,
                    "CMP": cmp,
                    "ScripName": tsym,
                    "OptionType": row['instrument_type'],
                    "StrikePrice": row['strike'],
                    "BidPrice": (q or {}).get('bid_price', 0),
                    "BidQty": (q or {}).get('bid_qty', 0),
                    "AskPrice": (q or {}).get('ask_price', 0),
                    "AskQty": (q or {}).get('ask_qty', 0),
                    "LTP": (q or {}).get('ltp', 0),
                    "Lot_Size": lot_size,
                    "Margin_Required": margin_val,
                    "LastUpdate": datetime.now().strftime("%H:%M:%S"),
                    "Days": days_to_expiry
                }
            
            return None
        
        except Exception:
            return None
    
    # Strategy 4: Per-symbol worker tuning based on contract count
    # More contracts = more workers, fewer contracts = fewer workers (less overhead)
    num_contracts = len(contracts_list)
    if num_contracts <= 5:
        optimal_workers = 2  # Small contract lists
    elif num_contracts <= 15:
        optimal_workers = 4  # Medium contract lists
    elif num_contracts <= 30:
        optimal_workers = 8  # Large contract lists
    else:
        optimal_workers = 12  # Very large contract lists
    
    # Cap by available workers
    max_workers = min(optimal_workers, current_option_workers * 2)
    
    # Use ThreadPoolExecutor with optimized concurrency
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(process_contract_balanced, contract) for contract in contracts_list]
        
        for future in as_completed(futures, timeout=120):
            try:
                result = future.result(timeout=15)
                if result:
                    results.append(result)
            except Exception:
                pass
    
    return results

def fetch_option_chain_zerodha_with_distance(symbol, cmp, distance_percentage):
    """Fetch option chain with distance-based filtering for live monitor"""
    global selected_expiry_global, symbol_matcher
    
    symbol_norm = symbol.strip().upper()
    
    # Use symbol matcher
    if symbol_matcher:
        matched_symbol = symbol_matcher.find_best_match(symbol_norm)
        if matched_symbol:
            symbol_norm = matched_symbol
    
    # Validate F&O availability
    if not validate_fo_availability(symbol_norm):
        return None
    
    instruments = fetch_instruments()
    if not instruments or 'nfo_options' not in instruments:
        return None
    
    nfo_options = instruments['nfo_options']
    
    # Filter options by symbol and expiry
    symbol_options = nfo_options[
        (nfo_options['name'].str.contains(f'\\b{symbol_norm}\\b', case=False, na=False, regex=True)) &
        (nfo_options['expiry'] == selected_expiry_global)
    ]
    
    # Fallback: exact match
    if len(symbol_options) == 0:
        symbol_options = nfo_options[
            (nfo_options['name'].str.contains(f'^{symbol_norm}$', case=False, na=False, regex=True)) &
            (nfo_options['expiry'] == selected_expiry_global)
        ]
    
    if len(symbol_options) == 0:
        return None
    
    # Apply distance-based strike selection logic
    distance_factor = distance_percentage / 100.0  # Convert percentage to decimal
    delta = cmp * distance_factor
    lower_bound = cmp - delta
    upper_bound = cmp + delta
    
    selected_contracts = symbol_options[
        (
            (symbol_options['instrument_type'] == 'PE') &
            (symbol_options['strike'] < lower_bound)
        ) |
        (
            (symbol_options['instrument_type'] == 'CE') &
            (symbol_options['strike'] > upper_bound)
        )
    ]
    
    if len(selected_contracts) == 0:
        return None
    
    results = []
    contracts_list = selected_contracts.to_dict('records')
    
    # Try batch processing first (async preferred)
    symbols_list = [row['tradingsymbol'] for row in contracts_list]
    quotes = {}
    if USE_ASYNC_FETCH and AIOHTTP_AVAILABLE:
        try:
            quotes = asyncio.run(fetch_quotes_batch_nfo_async(symbols_list))
        except Exception:
            quotes = {}
    if not quotes:
        try:
            quotes = fetch_quotes_batch_nfo(symbols_list)
        except Exception:
            quotes = {}
    
    def process_contract_live(row):
        """Process contract for live monitor with calculated columns"""
        try:
            tsym = row['tradingsymbol']
            lot_size = int(row['lot_size'])
            
            # Try to use batch quote data first
            q = quotes.get(tsym)
            
            # If batch data is missing, fetch individually
            if not q or (q.get('ltp', 0) == 0 and q.get('bid_price', 0) == 0):
                individual_data = fetch_option_data_optimized(tsym, lot_size)
                q = individual_data
                margin_val = individual_data.get('margin_required', 0)
            else:
                # Use batch quote data but fetch margin individually
                margin_val = 0
                if q.get('ltp', 0) > 0 or q.get('bid_price', 0) > 0:
                    margin_val = fetch_single_margin_with_validation(tsym, lot_size)
            
            ltp = (q or {}).get('ltp', 0)
            bid = (q or {}).get('bid_price', 0)
            
            if ltp > 0 or bid > 0:
                days_to_expiry = calculate_days_to_expiry(selected_expiry_global)
                
                # Calculate all derived columns
                lots = int(8000000 / margin_val) if margin_val > 0 else 0
                premium_lot = lot_size * bid
                cost = 12.0
                net_prem = premium_lot - cost
                total_prem = net_prem * lots
                yield_val = (total_prem / days_to_expiry) * 30 if days_to_expiry > 0 else 0
                distance = cmp - row['strike']
                dist_pct = distance / cmp if cmp > 0 else 0
                prem_day = total_prem / days_to_expiry if days_to_expiry > 0 else 0
                prem_until_expiry = total_prem
                
                return {
                    "Symbol": symbol_norm,
                    "CMP": cmp,
                    "ScripName": tsym,
                    "OptionType": row['instrument_type'],
                    "StrikePrice": row['strike'],
                    "BidPrice": (q or {}).get('bid_price', 0),
                    "BidQty": (q or {}).get('bid_qty', 0),
                    "AskPrice": (q or {}).get('ask_price', 0),
                    "AskQty": (q or {}).get('ask_qty', 0),
                    "LTP": (q or {}).get('ltp', 0),
                    "Lot_Size": lot_size,
                    "Margin_Required": margin_val,
                    "LastUpdate": datetime.now().strftime("%H:%M:%S"),
                    "Days": days_to_expiry,
                    # Calculated columns
                    "Lots": lots,
                    "Premium_Lot": premium_lot,
                    "Cost": cost,
                    "Net_Prem": net_prem,
                    "Total_Prem": total_prem,
                    "Yield": yield_val,
                    "Distance": distance,
                    "Dist_Pct": dist_pct,
                    "Prem_Day": prem_day,
                    "Prem_Until_Expiry": prem_until_expiry
                }
            
            return None
        
        except Exception:
            return None
    
    # Process contracts
    for contract in contracts_list:
        result = process_contract_live(contract)
        if result:
            results.append(result)
    
    return results

def fetch_option_chain_zerodha_live_monitor(symbol, cmp, distance_percentage, expiry_date):
    """Dedicated function for live monitor - doesn't affect global expiry"""
    global symbol_matcher
    
    symbol_norm = symbol.strip().upper()
    
    # Use symbol matcher
    if symbol_matcher:
        matched_symbol = symbol_matcher.find_best_match(symbol_norm)
        if matched_symbol:
            symbol_norm = matched_symbol
    
    # Validate F&O availability
    if not validate_fo_availability(symbol_norm):
        return None
    
    instruments = fetch_instruments()
    if not instruments or 'nfo_options' not in instruments:
        return None
    
    nfo_options = instruments['nfo_options']
    
    # Filter options by symbol and specific expiry (not global)
    symbol_options = nfo_options[
        (nfo_options['name'].str.contains(f'\\b{symbol_norm}\\b', case=False, na=False, regex=True)) &
        (nfo_options['expiry'] == expiry_date)
    ]
    
    if len(symbol_options) == 0:
        return None
    
    # Apply distance-based strike selection logic
    distance_factor = distance_percentage / 100.0
    delta = cmp * distance_factor
    lower_bound = cmp - delta
    upper_bound = cmp + delta
    
    selected_contracts = symbol_options[
        (
            (symbol_options['instrument_type'] == 'PE') &
            (symbol_options['strike'] < lower_bound)
        ) |
        (
            (symbol_options['instrument_type'] == 'CE') &
            (symbol_options['strike'] > upper_bound)
        )
    ]
    
    if len(selected_contracts) == 0:
        return None
    
    results = []
    contracts_list = selected_contracts.to_dict('records')
    
    # Batch fetch quotes
    symbols_list = [row['tradingsymbol'] for row in contracts_list]
    quotes = fetch_quotes_batch_nfo(symbols_list)
    
    # Process contracts with calculated columns
    for row in contracts_list:
        tsym = row['tradingsymbol']
        lot_size = int(row['lot_size'])
        
        q = quotes.get(tsym, {})
        
        # Fetch margin if quote has value
        margin_val = 0
        if q.get('ltp', 0) > 0 or q.get('bid_price', 0) > 0:
            margin_val = fetch_single_margin_with_validation(tsym, lot_size)
        
        ltp = q.get('ltp', 0)
        bid = q.get('bid_price', 0)
        
        if ltp > 0 or bid > 0:
            days_to_expiry = calculate_days_to_expiry(expiry_date)
            
            # Calculate all derived columns
            lots = int(8000000 / margin_val) if margin_val > 0 else 0
            premium_lot = lot_size * bid
            cost = 12.0
            net_prem = premium_lot - cost
            total_prem = net_prem * lots
            yield_val = (total_prem / days_to_expiry) * 30 if days_to_expiry > 0 else 0
            distance = cmp - row['strike']
            dist_pct = distance / cmp if cmp > 0 else 0
            prem_day = total_prem / days_to_expiry if days_to_expiry > 0 else 0
            prem_until_expiry = total_prem
            
            results.append({
                "Symbol": symbol_norm,
                "CMP": cmp,
                "ScripName": tsym,
                "OptionType": row['instrument_type'],
                "StrikePrice": row['strike'],
                "BidPrice": q.get('bid_price', 0),
                "BidQty": q.get('bid_qty', 0),
                "AskPrice": q.get('ask_price', 0),
                "AskQty": q.get('ask_qty', 0),
                "LTP": q.get('ltp', 0),
                "Lot_Size": lot_size,
                "Margin_Required": margin_val,
                "LastUpdate": datetime.now().strftime("%H:%M:%S"),
                "Days": days_to_expiry,
                # Calculated columns
                "Lots (Ill get)": lots,
                "Premium/Lot": premium_lot,
                "Cost": cost,
                "Net Prem": net_prem,
                "Total Prem": total_prem,
                "Yield": yield_val,
                "Distance": distance,
                "Dist %": dist_pct,
                "Prem/Day": prem_day,
                "Prem Until Expiry": prem_until_expiry
            })
    
    return results
    
def detect_existing_structure(filename, sheet_name):
    """Optimized Excel structure detection"""
    # Check cache first
    cached_data = excel_manager.get_cached_structure(filename, sheet_name)
    if cached_data:
        return cached_data
    
    try:
        wb = excel_manager.get_workbook(filename)
        if not wb or sheet_name not in wb.sheetnames:
            return None, None, None
        
        ws = wb[sheet_name]
        
        # Find data boundaries
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row <= 1:
            return None, None, None
        
        # Identify script data columns vs user formula columns
        script_data_cols = []
        user_formula_cols = []
        preserved_ranges = []
        
        # Check each column
        for col in range(1, max_col + 1):
            header_cell = ws.cell(row=1, column=col)
            header_value = str(header_cell.value) if header_cell.value else ""
            
            # Check if this is one of our script columns
            if header_value in SCRIPT_DATA_COLUMNS:
                script_data_cols.append(col)
            else:
                # This is a user-added column - preserve it
                user_formula_cols.append(col)
                
                # Check if it contains formulas (optimized sampling)
                has_formulas = False
                sample_cells = min(3, max_row - 1)  # Check only first 3 data rows
                for row in range(2, min(5, max_row + 1)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and str(cell.value).startswith('='):
                        has_formulas = True
                        break
                
                if has_formulas:
                    preserved_ranges.append({
                        'column': col,
                        'header': header_value,
                        'type': 'formula',
                        'range': f"{get_column_letter(col)}1:{get_column_letter(col)}{max_row}"
                    })
                elif header_value:
                    preserved_ranges.append({
                        'column': col,
                        'header': header_value,
                        'type': 'data',
                        'range': f"{get_column_letter(col)}1:{get_column_letter(col)}{max_row}"
                    })
        
        result = (script_data_cols, user_formula_cols, preserved_ranges)
        
        # Cache the result
        excel_manager.cache_structure(filename, sheet_name, result)
        
        return result
    
    except Exception:
        return None, None, None

def save_data_to_excel_with_preservation(all_data, filename):
    """Redirect saving to SQLite for performance while preserving workflow."""
    global _current_data_cache, _initial_processing_complete
    try:
        # Ensure the directory for the target file exists
        target_dir = os.path.dirname(filename) if os.path.dirname(filename) else OUTPUT_DIR
        if not os.path.exists(target_dir):
            os.makedirs(target_dir, exist_ok=True)
            safe_print(f"📁 Created directory: {target_dir}")
        
        safe_print(f"💾 Saving {len(all_data)} symbols to SQLite: {filename}")
        
        # Map to per-symbol tables, preserving current sheet-per-symbol structure
        # FIXED: Use the provided filename parameter instead of hardcoded OUTPUT_DB_FILE
        write_frames_as_tables(str(filename), all_data)

        # Update cache safely
        with data_collection_lock:
            _current_data_cache = all_data.copy()
            _initial_processing_complete = True
        
        # Verify file was created
        if os.path.exists(filename):
            file_size = os.path.getsize(filename)
            safe_print(f"🎉 Saved {len(all_data)} symbols to {filename} ({file_size:,} bytes)")
            return True
        else:
            safe_print(f"❌ File was not created: {filename}")
            return False
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        safe_print(f"❌ SQLite save error: {e}")
        safe_print(f"📋 Error details: {error_details}")
        return False

def cleanup_memory():
    """Clean up memory and caches"""
    global _current_data_cache, _initial_processing_complete
    
    # Only clear expired entries from caches
    cmp_cache.clear_expired_only()
    
    # Don't clear current data cache if initial processing isn't complete
    if _initial_processing_complete:
        excel_manager.clear_cache()
    
    # Force garbage collection
    gc.collect()

def optimize_for_next_cycle():
    """Optimize system for next cycle"""
    # Reset rate limiter
    rate_limiter.current_calls_per_second = rate_limiter.base_calls_per_second
    rate_limiter.timeout_count = 0
    
    # Reset circuit breaker
    circuit_breaker.failure_count = 0
    circuit_breaker.timeout_count = 0
    circuit_breaker.is_recovering = False
    
    # Clean up memory
    cleanup_memory()

def process_symbol(symbol, failed_symbols, all_data, current_workers):
    """Optimized symbol processing"""
    symbol_norm = symbol.strip().upper()
    
    if symbol_norm in INVALID_OPTION_SYMBOLS:
        return current_workers
    
    # Recovery check
    new_worker_count = circuit_breaker.check_and_recover_if_needed(rate_limiter, current_workers)
    if new_worker_count != current_workers:
        current_workers = new_worker_count
    
    # Fetch CMP
    cmp = fetch_cmp_zerodha(symbol_norm)
    if cmp is None:
        with data_collection_lock:
            failed_symbols.append(symbol_norm)
        return current_workers
    
    # Fetch option chain
    current_option_workers = min(current_workers, MAX_OPTION_WORKERS_PER_SYMBOL)
    option_data = fetch_option_chain_zerodha(symbol_norm, cmp, current_option_workers)
    
    if not option_data or len(option_data) == 0:
        with data_collection_lock:
            failed_symbols.append(symbol_norm)
        return current_workers
    
    # Save data
    df_option = pd.DataFrame(option_data)
    with data_collection_lock:
        all_data[symbol_norm] = df_option
    
    return current_workers

class RealTimeExcelUpdater:
    """Optimized real-time Excel updater"""
    
    def __init__(self):
        self.update_thread = None
        self.stop_updates = False
        self.last_update_time = 0
        self.wb = None
    
    def start_updates(self):
        """Start real-time updates after initial processing"""
        self.update_thread = threading.Thread(target=self._update_loop, daemon=True)
        self.update_thread.start()
    
    def _update_loop(self):
        """Update loop that waits for initial processing"""
        com_initialized = False
        
        try:
            # Wait for initial processing to complete
            while not _initial_processing_complete:
                time.sleep(10)
            
            # Initialize COM if available
            if XLWINGS_AVAILABLE and PYTHONCOM_AVAILABLE:
                try:
                    pythoncom.CoInitialize()
                    com_initialized = True
                except Exception:
                    pass
            
            # Connect to Excel if available
            if XLWINGS_AVAILABLE and os.path.exists(OUTPUT_EXCEL_FILE):
                try:
                    self.wb = xw.Book(OUTPUT_EXCEL_FILE)
                except Exception:
                    self.wb = None
            
            while not self.stop_updates:
                try:
                    current_time = time.time()
                    if current_time - self.last_update_time >= REAL_TIME_UPDATE_INTERVAL:
                        if _initial_processing_complete and _current_data_cache:
                            self._perform_real_update()
                            self.last_update_time = current_time
                    
                    time.sleep(10)
                
                except Exception:
                    time.sleep(30)
        
        finally:
            if self.wb and XLWINGS_AVAILABLE:
                try:
                    self.wb.close()
                except Exception:
                    pass
            
            if com_initialized and PYTHONCOM_AVAILABLE:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass
    
    def _perform_real_update(self):
        """Real-time update with thread-safe iteration"""
        with data_collection_lock:
            cached_data_copy = _current_data_cache.copy()
        
        if not cached_data_copy:
            return
        
        updated_data = {}
        update_count = 0
        
        # Process symbols one by one
        for symbol, df in cached_data_copy.items():
            try:
                # Get fresh CMP
                fresh_cmp = fetch_cmp_zerodha(symbol)
                if not fresh_cmp:
                    continue
                
                # Update CMP in dataframe
                df_copy = df.copy()
                df_copy['CMP'] = fresh_cmp
                df_copy['LastUpdate'] = datetime.now().strftime("%H:%M:%S")
                
                # Update days to expiry
                days_to_expiry = calculate_days_to_expiry(selected_expiry_global)
                df_copy['Days'] = days_to_expiry
                
                # Update a few key contracts
                sample_size = min(3, len(df_copy))
                for idx in df_copy.head(sample_size).index:
                    try:
                        row = df_copy.loc[idx]
                        scrip_name = str(row['ScripName'])
                        
                        # Fetch fresh quotes
                        session = session_pool.get_session()
                        try:
                            rate_limiter.wait_if_needed()
                            quote_response = session.get(
                                f'{KITE_BASE_URL}/quote', 
                                params={'i': f'NFO:{scrip_name}'},
                                timeout=5
                            )
                            
                            if quote_response.status_code == 200:
                                quote_data = quote_response.json()['data']
                                if f'NFO:{scrip_name}' in quote_data:
                                    option_data = quote_data[f'NFO:{scrip_name}']
                                    df_copy.at[idx, 'LTP'] = option_data.get('last_price', 0)
                                    
                                    depth = option_data.get('depth', {})
                                    if 'buy' in depth and depth['buy']:
                                        df_copy.at[idx, 'BidPrice'] = depth['buy'][0].get('price', 0)
                                        df_copy.at[idx, 'BidQty'] = depth['buy'][0].get('quantity', 0)
                                    
                                    if 'sell' in depth and depth['sell']:
                                        df_copy.at[idx, 'AskPrice'] = depth['sell'][0].get('price', 0)
                                        df_copy.at[idx, 'AskQty'] = depth['sell'][0].get('quantity', 0)
                                    
                                    update_count += 1
                        finally:
                            session_pool.return_session(session)
                    
                    except Exception:
                        continue
                
                updated_data[symbol] = df_copy
            
            except Exception:
                continue
            
            time.sleep(0.3)
        
        # Update cache
        if updated_data:
            with data_collection_lock:
                _current_data_cache.update(updated_data)
    
    def stop(self):
        """Stop updates"""
        self.stop_updates = True
        if self.wb and XLWINGS_AVAILABLE:
            try:
                self.wb.close()
            except:
                pass

# ---------------- Real-time ticker manager (polling with WS-ready design) ----------------
class RealTimeTickerManager:
    """Maintain live data for monitored option contracts per symbol.
    Uses HTTP polling; can be extended to WebSocket when kiteconnect is available.
    """

    def __init__(self):
        self.symbol_to_contracts = {}  # symbol -> list of tradingsymbols (NFO contracts)
        self.live_data = {}  # symbol -> list of rows (contract dicts)
        self.lock = threading.Lock()
        self.thread = None
        self.stop_flag = threading.Event()
        self.refresh_interval = 300  # 5 minutes = 300 seconds

    def start(self, symbols):
        """Start monitoring given symbols using cached processed data."""
        global _current_data_cache
        
        with self.lock:
            if self.thread and self.thread.is_alive():
                self.stop_flag.set()
                self.thread.join(timeout=5)
            
            self.symbol_to_contracts = {}
            self.live_data = {}
            
            # Build contract lists from cached data
            with data_collection_lock:
                cached_data = _current_data_cache.copy()
            
            for symbol in symbols:
                symbol_norm = symbol.strip().upper()
                if symbol_norm in cached_data:
                    df = cached_data[symbol_norm]
                    contracts = df['ScripName'].tolist()
                    self.symbol_to_contracts[symbol_norm] = contracts
                    self.live_data[symbol_norm] = df.to_dict('records')
            
            if not self.symbol_to_contracts:
                safe_print("⚠️ No cached data found for monitoring")
                return []
            
            self.stop_flag.clear()
            self.thread = threading.Thread(target=self._monitor_loop, daemon=True)
            self.thread.start()
            
            started_symbols = list(self.symbol_to_contracts.keys())
            safe_print(f"🎯 Started monitoring {len(started_symbols)} symbols: {started_symbols}")
            return started_symbols

    def _monitor_loop(self):
        """Background monitoring loop"""
        while not self.stop_flag.is_set():
            try:
                # Update live data
                self._update_live_data()
                
                # Wait for next refresh
                self.stop_flag.wait(self.refresh_interval)
                
            except Exception as e:
                safe_print(f"❌ Monitor loop error: {e}")
                time.sleep(30)

    def _update_live_data(self):
        """Update live data for all monitored symbols"""
        with self.lock:
            for symbol, contracts in self.symbol_to_contracts.items():
                try:
                    # Get fresh CMP
                    fresh_cmp = fetch_cmp_zerodha(symbol)
                    if not fresh_cmp:
                        continue
                    
                    # Update live data
                    if symbol in self.live_data:
                        for contract in self.live_data[symbol]:
                            contract['CMP'] = fresh_cmp
                            contract['LastUpdate'] = datetime.now().strftime("%H:%M:%S")
                            
                            # Update days to expiry
                            days_to_expiry = calculate_days_to_expiry(selected_expiry_global)
                            contract['Days'] = days_to_expiry
                    
                    # Update a few key contracts with fresh quotes
                    sample_contracts = contracts[:3]  # Update first 3 contracts
                    for contract_symbol in sample_contracts:
                        try:
                            session = session_pool.get_session()
                            try:
                                rate_limiter.wait_if_needed()
                                quote_response = session.get(
                                    f'{KITE_BASE_URL}/quote', 
                                    params={'i': f'NFO:{contract_symbol}'},
                                    timeout=5
                                )
                                
                                if quote_response.status_code == 200:
                                    quote_data = quote_response.json()['data']
                                    if f'NFO:{contract_symbol}' in quote_data:
                                        option_data = quote_data[f'NFO:{contract_symbol}']
                                        
                                        # Update contract in live data
                                        for contract in self.live_data[symbol]:
                                            if contract['ScripName'] == contract_symbol:
                                                contract['LTP'] = option_data.get('last_price', 0)
                                                
                                                depth = option_data.get('depth', {})
                                                if 'buy' in depth and depth['buy']:
                                                    contract['BidPrice'] = depth['buy'][0].get('price', 0)
                                                    contract['BidQty'] = depth['buy'][0].get('quantity', 0)
                                                
                                                if 'sell' in depth and depth['sell']:
                                                    contract['AskPrice'] = depth['sell'][0].get('price', 0)
                                                    contract['AskQty'] = depth['sell'][0].get('quantity', 0)
                                                break
                            except Exception:
                                pass
                        finally:
                            session_pool.return_session(session)
                    
                    time.sleep(0.2)  # Small delay between symbols
                
                except Exception:
                    continue

    def get_data(self):
        """Get current live data"""
        with self.lock:
            return self.live_data.copy()

    def stop(self):
        """Stop monitoring"""
        self.stop_flag.set()
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=5)
        safe_print("🛑 Real-time updates stopped")

# Global ticker manager instance
ticker_manager = RealTimeTickerManager()

def main():
    """Optimized main function"""
    global _initial_processing_complete
    
    safe_print("🚀 OPTIMIZED ZERODHA OPTIONS PROCESSOR")
    safe_print("=" * 50)
    safe_print(f"⚙️ Configuration:")
    safe_print(f"   • API Rate Limit: {API_CALLS_PER_SECOND} calls/second")
    safe_print(f"   • Symbol Workers: {MAX_SYMBOL_WORKERS}")
    safe_print(f"   • Option Workers: {MAX_OPTION_WORKERS_PER_SYMBOL}")
    safe_print(f"   • Session Pool: {session_pool.pool_size} connections")
    safe_print(f"   • Real-time updates every {REAL_TIME_UPDATE_INTERVAL}s")
    safe_print(f"   • Output Directory: {OUTPUT_DIR}")
    safe_print("=" * 50)
    
    # Test authentication
    if not test_zerodha_authentication():
        return
    
    # Load instruments
    if not fetch_instruments():
        safe_print("❌ Failed to load instruments")
        return
    
    # Read symbols from SQLite (migrate from Excel if present)
    try:
        df_main = read_symbols_list(MAIN_LIST_DB)
        if df_main is None or df_main.empty:
            if os.path.exists(MAIN_LIST_PATH):
                try:
                    df_excel = pd.read_excel(MAIN_LIST_PATH, sheet_name=0)
                    if 'Symbol' in df_excel.columns:
                        upsert_symbols_list(MAIN_LIST_DB, df_excel[['Symbol']])
                        df_main = df_excel
                    else:
                        safe_print("❌ No 'Symbol' column in Main_List.xlsx")
                        return
                except Exception as e:
                    safe_print(f"❌ Error migrating {MAIN_LIST_PATH} to SQLite: {e}")
                    return
            else:
                safe_print(f"❌ Neither {MAIN_LIST_DB} nor {MAIN_LIST_PATH} with 'Symbol' found.")
                return
    except Exception as e:
        safe_print(f"❌ Error reading symbols from SQLite: {e}")
        return
    
    if 'Symbol' not in df_main.columns:
        safe_print(f"❌ No 'Symbol' column found")
        return
    
    # Filter symbols
    all_symbols = df_main['Symbol'].dropna().astype(str).unique()
    symbols = [sym for sym in all_symbols if sym.upper() not in INVALID_OPTION_SYMBOLS]
    
    safe_print(f"🎯 Processing {len(symbols)} valid symbols")
    
    if len(symbols) == 0:
        safe_print("❌ No valid symbols found")
        return
    
    # Get expiry
    if not get_global_expiry(symbols[0]):
        return
    
    # PHASE 1: INITIAL PROCESSING
    safe_print("\n📊 PHASE 1: INITIAL DATA PROCESSING")
    
    failed_symbols = []
    all_data = {}
    current_workers = MAX_SYMBOL_WORKERS
    batch_size = current_workers
    
    for i in range(0, len(symbols), batch_size):
        current_batch = symbols[i:i+batch_size]
        
        with ThreadPoolExecutor(max_workers=current_workers) as executor:
            futures = [executor.submit(process_symbol, sym, failed_symbols, all_data, current_workers) for sym in current_batch]
            
            for future in as_completed(futures):
                try:
                    returned_workers = future.result()
                    if returned_workers != current_workers:
                        current_workers = returned_workers
                        batch_size = current_workers
                except Exception:
                    pass
        
        safe_print(f"📊 Progress: {i+len(current_batch)}/{len(symbols)} processed")
    
    # Save initial data
    if all_data:
        save_data_to_excel_with_preservation(all_data, OUTPUT_EXCEL_FILE)
    
    # Retry failed symbols
    if failed_symbols:
        safe_print(f"🔄 Retrying {len(failed_symbols)} failed symbols...")
        for symbol in failed_symbols[:3]:  # Limit retry attempts
            try:
                cmp = fetch_cmp_zerodha(symbol)
                if cmp and validate_fo_availability(symbol):
                    option_data = fetch_option_chain_zerodha(symbol, cmp, 1)
                    if option_data:
                        df_option = pd.DataFrame(option_data)
                        
                        # Ensure Days column is included
                        if 'Days' not in df_option.columns:
                            days_to_expiry = calculate_days_to_expiry(selected_expiry_global)
                            df_option['Days'] = days_to_expiry
                        
                        all_data[symbol] = df_option
                        failed_symbols.remove(symbol)
            except Exception:
                pass
    
    # Final save
    if all_data:
        save_data_to_excel_with_preservation(all_data, OUTPUT_EXCEL_FILE)
    
    safe_print("\n✅ INITIAL PROCESSING COMPLETED")
    
    # PHASE 2: REAL-TIME UPDATES
    start_realtime = input("\n🔄 Start real-time updates? (y/n): ").lower().strip()
    
    if start_realtime in ['y', 'yes']:
        safe_print("\n📡 PHASE 2: REAL-TIME UPDATES")
        
        updater = RealTimeExcelUpdater()
        updater.start_updates()
        
        safe_print("✅ Real-time updates running")
        safe_print("   ⏰ Press Ctrl+C to stop")
        
        try:
            cycle_count = 0
            while True:
                time.sleep(30)
                cycle_count += 1
                
                # Optimize system every 5 cycles
                if cycle_count % 5 == 0:
                    optimize_for_next_cycle()
                    
                    # Show performance statistics
                    cache_stats = cmp_cache.get_stats()
                    pool_stats = session_pool.get_stats()
                    safe_print(f"📊 Stats - Cache: {cache_stats['hit_rate']:.1f}%, Pool: {pool_stats['reused']}")
                    
        except KeyboardInterrupt:
            updater.stop()
            safe_print("\n🛑 Real-time updates stopped")
    
    # Final cleanup
    cleanup_memory()
    session_pool.close_all()
    
    # Final summary
    total = len(symbols)
    successful = len(all_data)
    failed = len(failed_symbols)
    
    safe_print("\n" + "=" * 50)
    safe_print("📊 FINAL SUMMARY:")
    safe_print(f"   • Total Symbols: {total}")
    safe_print(f"   • Successful: {successful}")
    safe_print(f"   • Failed: {failed}")
    safe_print(f"   • Success Rate: {(successful/total)*100:.1f}%")
    safe_print(f"   • Output DB: {OUTPUT_DB_FILE}")
    safe_print("=" * 50)
    
    if failed_symbols:
        safe_print("⚠️ Failed symbols:")
        for symbol in failed_symbols:
            fo_available = validate_fo_availability(symbol)
            status = "F&O not available" if not fo_available else "Other error"
            safe_print(f"   - {symbol}: {status}")
    
    safe_print("🎉 PROCESSING COMPLETED!")

if __name__ == "__main__":
    main()
