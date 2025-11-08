from app.main.live_monitor_manager import get_monitor_manager
from flask_login import current_user, login_required
from flask import render_template, request, redirect, url_for, jsonify, flash, session, send_file, abort
import importlib
import json
import pandas as pd
from datetime import datetime
import threading
import time
import os
from pathlib import Path
from . import main_bp

# Import enhanced database utilities
try:
    from .db_utils import (
        sqlite_connection, read_table, read_all_tables, list_tables,
        table_exists, get_database_size, create_database_backup,
        export_table_to_csv
    )
except ImportError:
    from app.main.db_utils import (
        sqlite_connection, read_table, read_all_tables, list_tables,
        table_exists, get_database_size, create_database_backup,
        export_table_to_csv
    )

# Global variables for managing background tasks
processing_status = {
    'is_running': False,
    'progress': 0,
    'total_symbols': 0,
    'current_symbol': '',
    'results': {},
    'failed_symbols': [],
    'start_time': None,
    'messages': []
}

# Filter job status
filter_status = {
    'is_running': False,
    'progress': 0,
    'start_time': None,
    'messages': [],
    'output_file': None,
    'summary': None,
    'tables': None
}

@main_bp.route("/")
def index():
    return render_template("index.html", title="Stock Options Dashboard")

@main_bp.route("/dashboard")
def dashboard():
    """Main dashboard showing all available functionalities"""
    yield_module = importlib.import_module('app.main.yield')
    
    # Check authentication status
    auth_status = yield_module.test_zerodha_authentication()
    
    # Get database statistics
    db_stats = {}
    
    # Check Options_Data.db
    options_db = "Processed_Options/Options_Data.db"
    if Path(options_db).exists():
        try:
            db_stats['options'] = get_database_size(options_db)
        except Exception as e:
            db_stats['options'] = {'exists': True, 'error': str(e)}
    
    # Check filter_output.db
    import glob
    filter_dbs = sorted(glob.glob('Filtered_Options/filter_output_*.db'), key=os.path.getmtime, reverse=True)
    if filter_dbs:
        try:
            db_stats['filtered'] = get_database_size(filter_dbs[0])
            db_stats['filtered']['path'] = filter_dbs[0]
        except Exception as e:
            db_stats['filtered'] = {'exists': True, 'error': str(e)}
    
    return render_template("dashboard.html", 
                         title="Dashboard", 
                         auth_status=auth_status,
                         processing_status=processing_status,
                         db_stats=db_stats)

@main_bp.route("/filter", methods=["GET", "POST"])
def filter_page():
    """Run and view filter processing on Options_Data.db"""
    if request.method == 'POST':
        # Start filter processing in background
        start_filter_background()
        # Stay on the filter page so UI can poll and render results
        return redirect(url_for('main.filter_page'))
    return render_template("filter.html", title="Filter Processor", status=filter_status)

def start_filter_background():
    global filter_status
    if filter_status['is_running']:
        return
    filter_status.update({
        'is_running': True,
        'progress': 0,
        'start_time': datetime.now(),
        'messages': ["Started filter processing on Options_Data.db"],
        'output_file': None,  # ✅ Correctly initialized to None - will be set to filter_output_YYYYMMDD_HHMMSS.db after processing completes
        'summary': None
    })
    thread = threading.Thread(target=run_filter_job, daemon=True)
    thread.start()

def run_filter_job():
    global filter_status
    try:
        yield_module = importlib.import_module('app.main.yield')
        filter_module = importlib.import_module('app.main.filter')
        
        # Check if Processed_Options/Options_Data.db exists
        processed_file = "Processed_Options/Options_Data.db"
        if not os.path.exists(processed_file):
            filter_status['messages'].append(f"❌ {processed_file} not found. Please run Process Options first.")
            filter_status.update({'is_running': False})
            return
        
        filter_status['messages'].append(f"📂 Starting filter processing workflow...")
        filter_status['messages'].append(f"📄 Using input file: {processed_file}")
        
        # Use the new optimized filter workflow
        result = filter_module.process_options_data_file_parallel(
            input_file=processed_file, 
            output_file=None, 
            max_workers=None
        )
        
        if result['success']:
            filter_status['summary'] = result
            # ✅ Set output_file from result - this is the timestamped filter_output_*.db file
            output_file_path = result.get('output_file')
            if output_file_path:
                filter_status['output_file'] = output_file_path
                filter_status['messages'].append(f"📁 Output file created: {os.path.basename(output_file_path)}")
            else:
                filter_status['messages'].append("⚠️ Warning: Output file path not returned from filter processing")
            
            filter_status['messages'].append("✅ Filter processing completed successfully!")
            filter_status['messages'].append(f"📊 Processed {result['processed_sheets']} sheets")
            filter_status['messages'].append(f"📈 Total rows: {result['total_rows']:,}")
            filter_status['messages'].append(f"🎯 Filtered records: {result['filtered_count']}")
            filter_status['messages'].append(f"🎯 High yield records: {result['high_yield_count']}")
            filter_status['messages'].append(f"⏱️ Total time: {result['total_time']:.2f}s")
            # Normalize summary keys for template compatibility
            try:
                if 'total_sheets' not in filter_status['summary']:
                    filter_status['summary']['total_sheets'] = filter_status['summary'].get('processed_sheets')
                if 'filtered_records' not in filter_status['summary']:
                    filter_status['summary']['filtered_records'] = filter_status['summary'].get('filtered_count', 0)
                if 'high_yield_records' not in filter_status['summary']:
                    filter_status['summary']['high_yield_records'] = filter_status['summary'].get('high_yield_count', 0)
            except Exception:
                pass
            
            # Load tables from the produced DB and prepare formatted display with conditional coloring
            try:
                try:
                    from .db_utils import read_all_tables
                except Exception:
                    from app.main.db_utils import read_all_tables
                
                # ✅ USE THE ACTUAL OUTPUT FILE PATH from result
                output_db_path = result.get('output_file')
                
                # Verify the file exists
                if not output_db_path or not os.path.exists(output_db_path):
                    # Fallback to latest timestamped file if the exact path doesn't exist
                    import glob
                    candidates = sorted(glob.glob('Filtered_Options/filter_output_*.db'), key=os.path.getmtime, reverse=True)
                    if candidates:
                        output_db_path = candidates[0]
                        filter_status['messages'].append(f"⚠️ Original output path not found, using latest: {output_db_path}")
                    else:
                        raise FileNotFoundError(f"Output file not found: {output_db_path}")
                
                filter_status['messages'].append(f"🗂️ Using output DB for display: {output_db_path}")
                
                # Verify file exists before reading
                if not os.path.exists(output_db_path):
                    raise FileNotFoundError(f"Database file does not exist: {output_db_path}")
                
                # Show file details for debugging
                import os.path as osp
                file_size = os.path.getsize(output_db_path)
                mod_time = datetime.fromtimestamp(os.path.getmtime(output_db_path)).strftime('%Y-%m-%d %H:%M:%S')
                filter_status['messages'].append(f"📦 File size: {file_size:,} bytes, Modified: {mod_time}")
                
                dfs = read_all_tables(output_db_path)
                
                # Get table names for debugging
                table_names = list(dfs.keys())
                filter_status['messages'].append(f"📑 Tables found in DB: {', '.join(sorted(table_names))}")
                filter_status['messages'].append(f"📊 Number of tables: {len(dfs)}")
                
                # Case-insensitive table access and debug available tables
                available = {str(k): k for k in dfs.keys()}
                lower_to_actual = {str(k).lower(): k for k in dfs.keys()}
                filter_status['messages'].append(f"🔍 Looking for tables: filtered, highyield")
                filter_status['messages'].append(f"🔍 Available (lowercase): {', '.join(sorted(lower_to_actual.keys()))}")
                tables = {}
                for canonical in ['filtered', 'highyield']:
                    actual_name = lower_to_actual.get(canonical)
                    if not actual_name:
                        filter_status['messages'].append(f"⚠️ Table '{canonical}' not found in database")
                        continue
                    df = dfs[actual_name].copy()
                    filter_status['messages'].append(f"✓ Found table '{actual_name}' with {len(df)} rows")
                    # Limit for UI
                    df = df.head(200)
                    # Prepare formatted rows and cell classes
                    columns = list(df.columns.astype(str))
                    formatted_rows = []
                    classes_rows = []
                    for _, row in df.iterrows():
                        formatted = {}
                        classes = {}
                        # number formatting
                        for col in columns:
                            val = row.get(col)
                            if pd.isna(val):
                                formatted[col] = ''
                                continue
                            # Format numbers with commas (but NOT BidPrice - display as-is)
                            if col in ['Yield', 'Prem/Day', 'Prem Until Expiry', 'Days Left', 'CMP', 'StrikePrice']:
                                try:
                                    formatted[col] = f"{float(val):,.0f}"
                                except Exception:
                                    formatted[col] = str(val)
                            elif col == 'BidPrice':
                                # Display BidPrice without formatting (as raw value)
                                try:
                                    formatted[col] = str(float(val))
                                except Exception:
                                    formatted[col] = str(val)
                            elif col == 'Dist %':
                                try:
                                    formatted[col] = f"{float(val):.2f}%"
                                except Exception:
                                    formatted[col] = str(val)
                            else:
                                formatted[col] = str(val)
                        # conditional coloring
                        try:
                            sheet_label = 'Filtered' if canonical == 'filtered' else 'HighYield'
                            if sheet_label == 'Filtered':
                                dist_val = row.get('Dist %')
                                dist_num = float(dist_val) if dist_val is not None and pd.notna(dist_val) else None
                                if dist_num is not None:
                                    if 15 <= dist_num < 18:
                                        cell_class = 'cell-light-red'
                                    elif 18 <= dist_num < 20:
                                        cell_class = 'cell-green'
                                    elif dist_num >= 20:
                                        cell_class = 'cell-light-green'
                                    else:
                                        cell_class = ''
                                    if cell_class:
                                        for c in ['Dist %', 'Prem/Day', 'Prem Until Expiry']:
                                            if c in columns:
                                                classes[c] = cell_class
                            elif sheet_label == 'HighYield':
                                yv = row.get('Yield')
                                ynum = float(yv) if yv is not None and pd.notna(yv) else None
                                if ynum is not None:
                                    if 80000 <= ynum < 100000:
                                        cell_class = 'cell-light-red'
                                    elif 100000 <= ynum < 150000:
                                        cell_class = 'cell-green'
                                    elif ynum >= 150000:
                                        cell_class = 'cell-light-green'
                                    else:
                                        cell_class = ''
                                    if cell_class:
                                        for c in ['Yield', 'Prem/Day', 'Prem Until Expiry']:
                                            if c in columns:
                                                classes[c] = cell_class
                        except Exception:
                            pass
                        formatted_rows.append(formatted)
                        classes_rows.append(classes)
                    tables[sheet_label] = {
                        'columns': columns,
                        'rows': formatted_rows,
                        'classes': classes_rows
                    }
                    filter_status['messages'].append(f"✓ Processed table '{sheet_label}': {len(formatted_rows)} rows formatted")
                
                if not tables:
                    filter_status['messages'].append(f"⚠️ No tables were loaded (tables dict is empty)")
                
                filter_status['tables'] = tables
                filter_status['display_file'] = output_db_path
                filter_status['messages'].append(f"📊 Filter results loaded from DB with formatting: {os.path.basename(output_db_path)}")
                filter_status['messages'].append(f"📋 Loaded {len(tables)} table(s) for display: {', '.join(tables.keys())}")
                for sheet, tbl_data in tables.items():
                    filter_status['messages'].append(f"  ✓ {sheet}: {len(tbl_data['rows'])} rows, {len(tbl_data['columns'])} columns")
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                filter_status['messages'].append(f"⚠️ Unable to load output DB tables: {e}")
                filter_status['messages'].append(f"Debug traceback: {error_details}")
        else:
            filter_status['messages'].append(f"❌ Filter processing failed: {result.get('error', 'Unknown error')}")
        
        filter_status.update({'is_running': False, 'progress': 100})
    except Exception as e:
        filter_status['messages'].append(f"❌ Filter error: {e}")
        filter_status.update({'is_running': False})

@main_bp.route("/api/filter-status")
def filter_status_api():
    return jsonify(filter_status)

@main_bp.route("/download/filter-output")
def download_filter_output():
    """
    Download filter output as Excel file with conditional formatting
    Converts the latest filter_output_*.db to Excel with color coding
    """
    try:
        import glob
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from io import BytesIO
        
        # Find the latest filter output DB
        db_candidates = sorted(glob.glob('Filtered_Options/filter_output_*.db'), key=os.path.getmtime, reverse=True)
        if not db_candidates:
            flash('No filter output file found. Please run the filter process first.', 'warning')
            return redirect(url_for('main.filter_page'))
        
        db_path = db_candidates[0]
        
        # Read tables from database
        dfs = read_all_tables(db_path)
        
        if not dfs:
            flash('Filter output database is empty.', 'warning')
            return redirect(url_for('main.filter_page'))
        
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Define color fills for conditional formatting
        light_red = PatternFill(start_color='FFB3BA', end_color='FFB3BA', fill_type='solid')
        green = PatternFill(start_color='BAE1A8', end_color='BAE1A8', fill_type='solid')
        light_green = PatternFill(start_color='7FD87F', end_color='7FD87F', fill_type='solid')
        
        # Process Filtered table
        if 'Filtered' in dfs:
            ws_filtered = wb.create_sheet('Filtered')
            df_filtered = dfs['Filtered']
            
            # Write headers
            headers = list(df_filtered.columns)
            ws_filtered.append(headers)
            
            # Style header row
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in ws_filtered[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write data rows
            for row_idx, (idx, row) in enumerate(df_filtered.iterrows(), start=2):
                row_data = []
                for col in headers:
                    val = row[col]
                    # Format numbers
                    if col in ['CMP', 'StrikePrice', 'BidPrice', 'Prem/Day', 'Prem Until Expiry', 'Yield']:
                        val = float(val) if pd.notna(val) else 0
                    elif col == 'Dist %':
                        val = float(val) if pd.notna(val) else 0
                    row_data.append(val)
                ws_filtered.append(row_data)
                
                # Apply conditional formatting based on Dist %
                dist_pct = float(row['Dist %']) if pd.notna(row.get('Dist %')) else 0
                row_num = row_idx  # row_idx already accounts for header row
                
                if 15 <= dist_pct < 18:
                    fill = light_red
                elif 18 <= dist_pct < 20:
                    fill = green
                elif dist_pct >= 20:
                    fill = light_green
                else:
                    fill = None
                
                if fill:
                    for cell in ws_filtered[row_num]:
                        cell.fill = fill
                        cell.font = Font(bold=True)
            
            # Format number columns
            for col_idx, col_name in enumerate(headers, start=1):
                col_letter = ws_filtered.cell(row=1, column=col_idx).column_letter
                if col_name in ['CMP', 'StrikePrice', 'BidPrice', 'Prem/Day', 'Prem Until Expiry', 'Yield']:
                    for row in ws_filtered.iter_rows(min_row=2, max_row=ws_filtered.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value is not None:
                                cell.number_format = '#,##0'
                elif col_name == 'Dist %':
                    for row in ws_filtered.iter_rows(min_row=2, max_row=ws_filtered.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value is not None:
                                cell.number_format = '0.00"%"'
            
            # Auto-adjust column widths
            for col in ws_filtered.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_filtered.column_dimensions[col_letter].width = adjusted_width
        
        # Process HighYield table
        if 'HighYield' in dfs:
            ws_highyield = wb.create_sheet('HighYield')
            df_highyield = dfs['HighYield']
            
            # Write headers
            headers = list(df_highyield.columns)
            ws_highyield.append(headers)
            
            # Style header row
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in ws_highyield[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write data rows
            for row_idx, (idx, row) in enumerate(df_highyield.iterrows(), start=2):
                row_data = []
                for col in headers:
                    val = row[col]
                    # Format numbers
                    if col in ['CMP', 'StrikePrice', 'BidPrice', 'Prem/Day', 'Prem Until Expiry', 'Yield']:
                        val = float(val) if pd.notna(val) else 0
                    elif col == 'Dist %':
                        val = float(val) if pd.notna(val) else 0
                    row_data.append(val)
                ws_highyield.append(row_data)
                
                # Apply conditional formatting based on Yield
                yield_val = float(row['Yield']) if pd.notna(row.get('Yield')) else 0
                row_num = row_idx  # row_idx already accounts for header row
                
                if 80000 <= yield_val < 100000:
                    fill = light_red
                elif 100000 <= yield_val < 150000:
                    fill = green
                elif yield_val >= 150000:
                    fill = light_green
                else:
                    fill = None
                
                if fill:
                    for cell in ws_highyield[row_num]:
                        cell.fill = fill
                        cell.font = Font(bold=True)
            
            # Format number columns
            for col_idx, col_name in enumerate(headers, start=1):
                col_letter = ws_highyield.cell(row=1, column=col_idx).column_letter
                if col_name in ['CMP', 'StrikePrice', 'BidPrice', 'Prem/Day', 'Prem Until Expiry', 'Yield']:
                    for row in ws_highyield.iter_rows(min_row=2, max_row=ws_highyield.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value is not None:
                                cell.number_format = '#,##0'
                elif col_name == 'Dist %':
                    for row in ws_highyield.iter_rows(min_row=2, max_row=ws_highyield.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value is not None:
                                cell.number_format = '0.00"%"'
            
            # Auto-adjust column widths
            for col in ws_highyield.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_highyield.column_dimensions[col_letter].width = adjusted_width
        
        # Save to BytesIO buffer (in-memory, no temp files)
        output = BytesIO()
        try:
            wb.save(output)
            output.seek(0)
            
            # Generate filename with timestamp
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f'filter_output_{timestamp}.xlsx'
            
            # Close workbook to free resources
            wb.close()
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        finally:
            # Ensure workbook is closed even if send_file fails
            try:
                wb.close()
            except:
                pass
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Error generating Excel file: {str(e)}', 'error')
        return redirect(url_for('main.filter_page'))

def get_global_expiry_for_symbol(symbol: str):
    """Return sorted expiry dates (as strings) for the given symbol using app.main.yield."""
    yield_module = importlib.import_module('app.main.yield')
    instruments = yield_module.fetch_instruments()
    
    if not instruments or 'nfo_options' not in instruments:
        return []
        
    nfo_options = instruments['nfo_options']
    
    # Filter by symbol and get unique expiry dates
    filtered_df = nfo_options[nfo_options['name'].str.contains(symbol, case=False, na=False)]
    unique_expiry_values = sorted(filtered_df['expiry'].unique())
    
    return unique_expiry_values

@main_bp.route("/expiry", methods=["GET", "POST"])
def select_expiry():
    """Expiry selection page with support for multiple symbols"""
    yield_module = importlib.import_module('app.main.yield')
    
    # Test authentication
    if not yield_module.test_zerodha_authentication():
        return redirect(url_for('main.auth'))
    
    # Get symbol from form or default
    symbol = request.form.get('symbol', 'RELIANCE').upper()
    
    if request.method == 'POST':
        selected_date = request.form.get('expiry')
        if selected_date:
            # Store selected expiry in session
            session['selected_expiry'] = selected_date
            session['selected_symbol'] = symbol
            flash(f'Selected expiry: {selected_date} for {symbol}', 'success')
            return redirect(url_for('main.process_options'))
    
    expiry_dates = get_global_expiry_for_symbol(symbol)
    return render_template("expiry.html", 
                         title="Select Expiry", 
                         symbol=symbol, 
                         expiry_dates=expiry_dates,
                         selected_date=session.get('selected_expiry'))

@main_bp.route("/process-options", methods=["GET", "POST"])
def process_options():
    """Process options for selected symbols and expiry"""
    yield_module = importlib.import_module('app.main.yield')
    
    # Check authentication
    if not yield_module.test_zerodha_authentication():
        return redirect(url_for('main.auth'))
    
    if request.method == 'POST':
        symbols_input = request.form.get('symbols', '').strip()
        selected_expiry = session.get('selected_expiry')
        
        if not selected_expiry:
            flash('Please select an expiry date first', 'error')
            return redirect(url_for('main.select_expiry'))
        
        symbols: list[str] = []
        # Parse symbols from textarea
        if symbols_input:
            symbols.extend([s.strip().upper() for s in symbols_input.split(',') if s.strip()])
        
        # Excel uploads are no longer supported in DB-only mode
        uploaded = request.files.get('symbols_file')
        if uploaded and uploaded.filename:
            flash('Excel uploads are not supported. Please enter symbols manually or load from DB.', 'error')
            return redirect(url_for('main.process_options'))
        
        # Deduplicate and filter empties
        symbols = sorted(list({s for s in symbols if s}))
        
        if not symbols:
            flash('Please provide at least one symbol (textarea or Excel)', 'error')
            return redirect(url_for('main.process_options'))
        
        # Start background processing
        start_background_processing(symbols, selected_expiry)
        
        return redirect(url_for('main.processing_status_page'))
    
    return render_template("process_options.html", 
                         title="Process Options",
                         selected_expiry=session.get('selected_expiry'),
                         selected_symbol=session.get('selected_symbol', 'RELIANCE'))

def start_background_processing(symbols, selected_expiry):
    """Start background processing of options data"""
    global processing_status
    
    # Always start fresh processing when user clicks "Start Processing"
    # Reset status
    processing_status.update({
        'is_running': True,
        'progress': 0,
        'total_symbols': len(symbols),
        'current_symbol': '',
        'results': {},
        'failed_symbols': [],
        'start_time': datetime.now(),
        'messages': [f'🚀 Starting fresh processing of {len(symbols)} symbols with expiry {selected_expiry}']
    })
    
    # Start processing in background thread
    thread = threading.Thread(target=process_symbols_background, args=(symbols, selected_expiry))
    thread.daemon = True
    thread.start()

def process_symbols_background(symbols, selected_expiry):
    """Background thread function for processing symbols using optimized approach"""
    global processing_status
    yield_module = importlib.import_module('app.main.yield')
    
    try:
        # Set global expiry in yield module
        yield_module.selected_expiry_global = selected_expiry
        
        processing_status['messages'].append(f'🚀 Starting optimized processing of {len(symbols)} symbols...')
        
        # Ensure Processed_Options directory exists
        processed_options_dir = "Processed_Options"
        if not os.path.exists(processed_options_dir):
            os.makedirs(processed_options_dir)
            processing_status['messages'].append(f'📁 Created {processed_options_dir} directory')
        
        # Backup existing DB file if it exists
        db_file = os.path.join(processed_options_dir, "Options_Data.db")
        if os.path.exists(db_file):
            backup_file = os.path.join(processed_options_dir, f"Options_Data_backup_{int(time.time())}.db")
            try:
                import shutil
                shutil.copy2(db_file, backup_file)
                processing_status['messages'].append(f'📁 Backed up previous results to {backup_file}')
            except Exception as e:
                processing_status['messages'].append(f'⚠️ Could not backup previous file: {e}')
        
        # Use the optimized processing approach
        failed_symbols = []
        all_data = {}
        current_workers = getattr(yield_module, 'MAX_SYMBOL_WORKERS', 12)  # Use optimized default
        processed_count = 0
        
        # Process symbols using the optimized process_symbol function
        batch_size = current_workers
        total_symbols = len(symbols)
        
        for i in range(0, total_symbols, batch_size):
            current_batch = symbols[i:i+batch_size]
            batch_start = i + 1
            batch_end = min(i + batch_size, total_symbols)
            
            processing_status['current_symbol'] = f'Batch {batch_start}-{batch_end}'
            processing_status['progress'] = int((i / total_symbols) * 100)
            processing_status['messages'].append(f'📊 Processing batch {batch_start}-{batch_end} of {total_symbols} symbols...')
            
            # Process batch using optimized approach with ThreadPoolExecutor
            from concurrent.futures import ThreadPoolExecutor, as_completed
            
            with ThreadPoolExecutor(max_workers=current_workers) as executor:
                futures = [executor.submit(yield_module.process_symbol, sym, failed_symbols, all_data, current_workers) for sym in current_batch]
                
                completed_in_batch = 0
                for future in as_completed(futures):
                    try:
                        returned_workers = future.result()
                        if returned_workers != current_workers:
                            current_workers = returned_workers
                            batch_size = current_workers
                        
                        completed_in_batch += 1
                        processed_count += 1
                        
                        # Update progress for each completed symbol
                        if total_symbols > 0:
                            processing_status['progress'] = int((processed_count / total_symbols) * 100)
                            # Update current symbol being processed in X/Y format
                            processing_status['current_symbol'] = f'{processed_count}/{total_symbols}'
                        
                    except Exception as e:
                        processing_status['messages'].append(f'⚠️ Worker error: {str(e)}')
                        processed_count += 1
                        if total_symbols > 0:
                            processing_status['progress'] = int((processed_count / total_symbols) * 100)
                            processing_status['current_symbol'] = f'{processed_count}/{total_symbols}'
            
            processing_status['messages'].append(f'✅ Batch {batch_start}-{batch_end} completed: {completed_in_batch}/{len(current_batch)} successful')
        
        # Update progress to 90% while saving
        processing_status['progress'] = 90
        processing_status['current_symbol'] = 'Saving data to database...'
        
        # Save to SQLite in Processed_Options directory
        if all_data:
            processing_status['messages'].append('💾 Saving fresh data to Processed_Options/Options_Data.db...')
            db_file = os.path.join(processed_options_dir, "Options_Data.db")
            
            # FIXED: Use temporary file approach to avoid Windows file locking issues
            # Write to temp file first, then atomically replace
            temp_db_file = db_file + '.tmp'
            
            try:
                # Close any existing connections to the old file
                from app.main.db_utils import wait_for_writes, shutdown_writers
                
                # Wait for any pending writes to complete
                wait_for_writes(db_file)
                
                # Write to temporary file first
                processing_status['messages'].append(f'📝 Writing to temporary file: {temp_db_file}')
                success = yield_module.save_data_to_excel_with_preservation(all_data, temp_db_file)
                
                if success and os.path.exists(temp_db_file):
                    processing_status['messages'].append(f'✅ Temporary file created successfully: {temp_db_file}')
                    # Small delay to ensure file is fully written and released
                    time.sleep(0.1)
                    
                    # Remove old file if it exists (with retry logic for Windows)
                    if os.path.exists(db_file):
                        max_retries = 5
                        for attempt in range(max_retries):
                            try:
                                # Try to close any lingering connections first
                                import sqlite3
                                try:
                                    test_conn = sqlite3.connect(db_file, timeout=1.0)
                                    test_conn.close()
                                except:
                                    pass
                                
                                os.remove(db_file)
                                break
                            except (OSError, PermissionError) as e:
                                if attempt < max_retries - 1:
                                    time.sleep(0.2 * (attempt + 1))  # Exponential backoff
                                else:
                                    # If we can't remove, try renaming it out of the way
                                    try:
                                        old_backup = db_file + '.old'
                                        if os.path.exists(old_backup):
                                            os.remove(old_backup)
                                        os.rename(db_file, old_backup)
                                        processing_status['messages'].append(f'📦 Renamed old file to backup')
                                    except Exception as rename_err:
                                        processing_status['messages'].append(f'⚠️ Could not remove/rename old file: {e}')
                    
                    # Atomically replace with temp file (works on Windows)
                    try:
                        os.replace(temp_db_file, db_file)
                        processing_status['messages'].append('✅ Fresh data saved to Processed_Options/Options_Data.db')
                        processing_status['messages'].append('ℹ️ Data fetching complete. Use Filter to add calculated columns.')
                    except Exception as replace_err:
                        # Fallback: copy temp file and remove temp
                        import shutil
                        try:
                            shutil.copy2(temp_db_file, db_file)
                            os.remove(temp_db_file)
                            processing_status['messages'].append('✅ Fresh data saved to Processed_Options/Options_Data.db')
                            processing_status['messages'].append('ℹ️ Data fetching complete. Use Filter to add calculated columns.')
                        except Exception as copy_err:
                            processing_status['messages'].append(f'❌ Failed to replace database file: {replace_err}')
                            if os.path.exists(temp_db_file):
                                os.remove(temp_db_file)
                else:
                    # Provide more detailed error information
                    error_msg = '❌ Failed to save database file'
                    if not success:
                        error_msg += ' (save function returned False)'
                    elif not os.path.exists(temp_db_file):
                        error_msg += f' (temp file not found: {temp_db_file})'
                    processing_status['messages'].append(error_msg)
                    
                    # Log additional debugging info
                    if os.path.exists(os.path.dirname(temp_db_file)):
                        processing_status['messages'].append(f'📁 Directory exists: {os.path.dirname(temp_db_file)}')
                    else:
                        processing_status['messages'].append(f'❌ Directory missing: {os.path.dirname(temp_db_file)}')
                    
                    if os.path.exists(temp_db_file):
                        try:
                            os.remove(temp_db_file)
                        except Exception as cleanup_err:
                            processing_status['messages'].append(f'⚠️ Could not remove temp file: {cleanup_err}')
            except Exception as e:
                processing_status['messages'].append(f'❌ Error saving database: {e}')
                if os.path.exists(temp_db_file):
                    try:
                        os.remove(temp_db_file)
                    except:
                        pass
        
        # Update final status
        processing_status.update({
            'is_running': False,
            'progress': 100,
            'current_symbol': '',
            'results': {symbol: df.to_dict('records') for symbol, df in all_data.items()},
            'failed_symbols': failed_symbols,
            'messages': processing_status['messages'] + [f'🎉 Processing completed! Success: {len(all_data)}, Failed: {len(failed_symbols)}']
        })
        
    except Exception as e:
        processing_status.update({
            'is_running': False,
            'progress': 0,
            'current_symbol': '',
            'messages': processing_status['messages'] + [f'❌ Fatal error: {str(e)}']
        })

@main_bp.route("/processing-status")
def processing_status_page():
    """Show processing status page"""
    return render_template("processing_status.html", 
                         title="Processing Status",
                         status=processing_status)

@main_bp.route("/api/processing-status")
def api_processing_status():
    """API endpoint for processing status (for AJAX updates)"""
    return jsonify(processing_status)

@main_bp.route("/results")
def view_results():
    """View processing results"""
    # Check for results in Processed_Options directory first
    processed_file = "Processed_Options/Options_Data.db"
    
    # If no results in memory but DB file exists, try to load from SQLite
    if not processing_status['results'] and os.path.exists(processed_file):
        try:
            try:
                from .db_utils import read_all_tables
            except Exception:
                from app.main.db_utils import read_all_tables
            all_tables = read_all_tables(processed_file)
            results = {name: df.to_dict('records') for name, df in all_tables.items()}
            processing_status['results'] = results
        except Exception as e:
            print(f"Error loading results from SQLite: {e}")
    
    return render_template("results.html", 
                         title="Processing Results",
                         results=processing_status['results'],
                         failed_symbols=processing_status['failed_symbols'])

@main_bp.route("/download/excel")
def download_excel():
    """Download the generated SQLite database if available."""
    # Check Processed_Options directory first
    processed_file = "Processed_Options/Options_Data.db"
    if os.path.exists(processed_file):
        try:
            return send_file(processed_file, as_attachment=True, download_name="Options_Data.db")
        except Exception:
            pass
    
    # Fallback to root directory
    file_path = "Options_Data.db"
    try:
        return send_file(file_path, as_attachment=True, download_name=file_path)
    except Exception:
        abort(404)

@main_bp.route("/symbol-lookup", methods=["GET", "POST"])
def symbol_lookup():
    """Symbol lookup and validation tool"""
    yield_module = importlib.import_module('app.main.yield')
    
    results = []
    if request.method == 'POST':
        symbols_input = request.form.get('symbols', '')
        symbols = [s.strip().upper() for s in symbols_input.split(',') if s.strip()]
        
        for symbol in symbols:
            try:
                # Check if instruments are loaded
                instruments = yield_module.fetch_instruments()
                if not instruments:
                    results.append({
                        'symbol': symbol,
                        'status': 'error',
                        'message': 'Failed to fetch instruments',
                        'cmp': None,
                        'fo_available': False
                    })
                    continue
                
                # Get CMP
                cmp = yield_module.fetch_cmp_zerodha(symbol)
                
                # Check F&O availability
                fo_available = yield_module.validate_fo_availability(symbol)
                
                results.append({
                    'symbol': symbol,
                    'status': 'success' if cmp else 'warning',
                    'message': 'Found' if cmp else 'Symbol not found or no price available',
                    'cmp': cmp,
                    'fo_available': fo_available
                })
                
            except Exception as e:
                results.append({
                    'symbol': symbol,
                    'status': 'error',
                    'message': f'Error: {str(e)}',
                    'cmp': None,
                    'fo_available': False
                })
    
    return render_template("symbol_lookup.html", 
                        title="Symbol Lookup",
                        results=results)


# ============================================================================
# LIVE MONITOR ROUTES (Per-User Isolated Sessions)
# ============================================================================

@main_bp.route("/live-monitor")
def live_monitor():
    """
    Enhanced live monitor - accessible to all
    Shows login prompt if not authenticated
    """
    return render_template("live_monitor.html", title="Live Options Monitor")

@main_bp.route("/api/live-monitor/debug")
def api_live_monitor_debug():
    """Debug endpoint to see what's happening in the monitor"""
    try:
        from app.main.live_monitor_manager import get_monitor_manager
        
        # Determine user ID
        user_id = current_user.id if current_user.is_authenticated else -1
        
        manager = get_monitor_manager()
        session = manager.get_session(user_id)
        
        if not session:
            return jsonify({'error': 'No active monitoring session'}), 404
        
        # Get detailed state
        with session.data_lock:
            debug_info = {
                'is_running': session.is_running,
                'user_id': session.user_id,
                'symbols': session.symbols,
                'expiry_date': session.expiry_date,
                'distance_percentage': session.distance_percentage,
                'last_update': session.last_update.isoformat() if session.last_update else None,
                'stats': session.stats,
                'raw_data_symbols': list(session.raw_data.keys()),
                'calculated_data_symbols': list(session.calculated_data.keys()),
                'filtered_data': {
                    'filtered_count': len(session.filtered_data.get('filtered', [])),
                    'high_yield_count': len(session.filtered_data.get('high_yield', [])),
                    'filtered_sample': session.filtered_data.get('filtered', [])[:3],  # First 3
                    'high_yield_sample': session.filtered_data.get('high_yield', [])[:3]  # First 3
                },
                'changes_queue_size': session.changes_queue.qsize(),
                'worker_thread_alive': session.worker_thread.is_alive() if session.worker_thread else False
            }
        
        return jsonify({
            'status': 'success',
            'debug': debug_info
        }), 200
        
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500
        
@main_bp.route("/api/live-monitor/start", methods=["POST"])
def api_live_monitor_start():
    """
    Start live monitoring
    Works with or without authentication (uses config credentials if not logged in)
    """
    try:
        payload = request.get_json(silent=True) or {}
        symbols = payload.get('symbols', [])
        distance_percentage = payload.get('distance_percentage', 10)
        expiry_date = payload.get('expiry_date')
        refresh_interval = payload.get('refresh_interval', 10)  # seconds
        
        # Validation
        if not symbols:
            return jsonify({'error': 'Symbols list is required'}), 400
        
        if not expiry_date:
            return jsonify({'error': 'Expiry date is required'}), 400
        
        # Determine user context
        if current_user.is_authenticated:
            # Use logged-in user's credentials
            user_id = current_user.id
            
            from app.user_runtime import get_user_runtime_context
            from app.models import User
            
            user = User.get_by_id(user_id)
            if not user:
                return jsonify({'error': 'User not found'}), 401
            
            # Get decrypted credentials
            api_key, access_token = user.get_credentials()
            if not api_key or not access_token:
                return jsonify({
                    'error': 'Please configure your Zerodha API credentials first',
                    'redirect': '/auth/credentials'
                }), 401
            
            # Get or create runtime context for this user
            runtime_ctx = get_user_runtime_context(user_id, api_key, access_token)
        else:
            # Use global config credentials (for non-authenticated users)
            # This allows testing without login
            yield_module = importlib.import_module('app.main.yield')
            api_key = yield_module.API_KEY
            access_token = yield_module.ACCESS_TOKEN
            
            if not api_key or not access_token:
                return jsonify({
                    'error': 'No credentials available. Please login or set ZERODHA_API_KEY and ZERODHA_ACCESS_TOKEN environment variables',
                    'redirect': '/auth/login'
                }), 401
            
            # Use a guest user ID (negative to avoid conflicts)
            user_id = -1
            
            from app.user_runtime import get_user_runtime_context
            runtime_ctx = get_user_runtime_context(user_id, api_key, access_token)
        
        # Create monitoring session
        from app.main.live_monitor_manager import get_monitor_manager
        
        manager = get_monitor_manager()
        session = manager.create_session(
            user_id=user_id,
            runtime_context=runtime_ctx,
            symbols=symbols,
            expiry_date=expiry_date,
            distance_percentage=distance_percentage
        )
        
        # Start the session
        success = session.start(refresh_interval=refresh_interval)
        
        if success:
            return jsonify({
                'status': 'started',
                'user_id': user_id if user_id > 0 else 'guest',
                'symbols': symbols,
                'expiry_date': expiry_date,
                'distance_percentage': distance_percentage,
                'refresh_interval': refresh_interval,
                'message': 'Live monitoring started successfully'
            }), 200
        else:
            return jsonify({'error': 'Failed to start monitoring session'}), 500
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@main_bp.route("/api/live-monitor/stop", methods=["POST"])
def api_live_monitor_stop():
    """Stop live monitoring (works for both authenticated and guest users)"""
    try:
        from app.main.live_monitor_manager import get_monitor_manager
        
        # Determine user ID
        user_id = current_user.id if current_user.is_authenticated else -1
        
        manager = get_monitor_manager()
        manager.stop_session(user_id)
        
        return jsonify({
            'status': 'stopped',
            'message': 'Live monitoring stopped'
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@main_bp.route("/api/live-monitor/data")
def api_live_monitor_data():
    """Get current filtered data (works for both authenticated and guest users)"""
    try:
        from app.main.live_monitor_manager import get_monitor_manager
        
        # Determine user ID
        user_id = current_user.id if current_user.is_authenticated else -1
        
        manager = get_monitor_manager()
        session = manager.get_session(user_id)
        
        if not session:
            return jsonify({'error': 'No active monitoring session'}), 404
        
        data = session.get_current_data()
        
        return jsonify({
            'status': 'success',
            'data': data,
            'timestamp': datetime.now().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@main_bp.route("/api/live-monitor/changes")
def api_live_monitor_changes():
    """Get incremental changes (works for both authenticated and guest users)"""
    try:
        from app.main.live_monitor_manager import get_monitor_manager
        
        # Determine user ID
        user_id = current_user.id if current_user.is_authenticated else -1
        
        manager = get_monitor_manager()
        session = manager.get_session(user_id)
        
        if not session:
            return jsonify({'error': 'No active monitoring session'}), 404
        
        # Wait for changes (with timeout)
        timeout = float(request.args.get('timeout', 30.0))
        changes = session.get_changes(timeout=timeout)
        
        if changes:
            return jsonify({
                'status': 'success',
                'changes': changes,
                'has_changes': True
            }), 200
        else:
            # Timeout - no changes
            return jsonify({
                'status': 'success',
                'has_changes': False,
                'message': 'No changes in timeout period'
            }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@main_bp.route("/api/live-monitor/status")
def api_live_monitor_status():
    """Get monitoring session status (works for both authenticated and guest users)"""
    try:
        from app.main.live_monitor_manager import get_monitor_manager
        
        # Determine user ID
        user_id = current_user.id if current_user.is_authenticated else -1
        
        manager = get_monitor_manager()
        session = manager.get_session(user_id)
        
        if not session:
            return jsonify({
                'is_running': False,
                'message': 'No active session'
            }), 200
        
        data = session.get_current_data()
        
        return jsonify({
            'is_running': data['is_running'],
            'last_update': data['last_update'],
            'stats': data['stats'],
            'symbols_count': data['symbols_count']
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@main_bp.route("/api/live-monitor/expiry-dates")
def api_live_monitor_expiry_dates():
    """Get available expiry dates (no auth required)"""
    try:
        yield_module = importlib.import_module('app.main.yield')
        
        # Fetch instruments
        instruments = yield_module.fetch_instruments()
        if not instruments or 'nfo_options' not in instruments:
            return jsonify({'error': 'Failed to fetch instruments'}), 500
        
        nfo_options = instruments['nfo_options']
        
        # Get unique expiry dates and sort them
        expiry_dates = sorted(nfo_options['expiry'].dropna().unique())
        
        # Convert to string format for frontend
        expiry_strings = [str(date) for date in expiry_dates]
        
        return jsonify({
            'status': 'success',
            'expiry_dates': expiry_strings,
            'count': len(expiry_strings)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _init_zerodha_session_pool(access_token_override: str | None = None):
    """Initialize a fresh Zerodha session pool and verify auth."""
    yield_module = importlib.import_module('app.main.yield')
    
    try:
        # Optionally override access token
        if access_token_override:
            yield_module.ACCESS_TOKEN = access_token_override
            
            # Note: Environment variables should be updated externally
            # We update the in-memory value for this session
            print(f"Updated access token for current session (environment variable unchanged)")
            
            # Also refresh headers on any existing pool sessions if present
            try:
                if hasattr(yield_module, 'session_pool') and yield_module.session_pool:
                    for s in list(getattr(yield_module.session_pool, 'pool', [])):
                        s.headers.update({'Authorization': f'token {yield_module.API_KEY}:{yield_module.ACCESS_TOKEN}'})
            except Exception as e:
                print(f"Warning: Could not update existing session headers: {e}")
        
        # Create a fresh session pool and assign to module-level
        new_pool = yield_module.OptimizedSessionPool()
        yield_module.session_pool = new_pool
        
        # Quick sanity check: create a single session
        _ = yield_module.create_zerodha_session()
        
        # Verify authentication via provided helper
        print(f"Testing authentication with API_KEY: {yield_module.API_KEY[:10]}... and ACCESS_TOKEN: {yield_module.ACCESS_TOKEN[:10]}...")
        ok = yield_module.test_zerodha_authentication()
        
        if not ok:
            return False, "Authentication test failed - please check your access token"
        
        return ok, None
        
    except AttributeError as e:
        return False, f"Missing attribute in yield module: {e}"
    except ImportError as e:
        return False, f"Failed to import yield module: {e}"
    except Exception as e:
        error_msg = str(e) if str(e) else f"Exception type: {type(e).__name__}"
        return False, f"Authentication error: {error_msg}"

@main_bp.route("/auth", methods=["GET", "POST"])
def auth():
    """Authentication page"""
    status = None
    message = None
    yield_module = importlib.import_module('app.main.yield')
    api_key = yield_module.API_KEY
    current_token = None
    
    if request.method == 'POST':
        submitted_token = request.form.get('access_token', '').strip() or None
        current_token = submitted_token
        
        ok, err = _init_zerodha_session_pool(submitted_token)
        
        if ok:
            flash('Authentication successful!', 'success')
            return redirect(url_for('main.dashboard'))
        
        status = 'error'
        message = f'Authentication failed: {err or "Unknown error"}'
    
    return render_template("auth.html", 
                         title="Authenticate", 
                         status=status, 
                         message=message, 
                         api_key=api_key, 
                         access_token=current_token)

@main_bp.route("/settings")
def settings():
    """Settings and configuration page"""
    yield_module = importlib.import_module('app.main.yield')
    
    config_data = {
        'API_KEY': yield_module.API_KEY,
        'KITE_BASE_URL': yield_module.KITE_BASE_URL,
        'MAX_SYMBOL_WORKERS': getattr(yield_module, 'MAX_SYMBOL_WORKERS', 8),
        'MAX_OPTION_WORKERS_PER_SYMBOL': getattr(yield_module, 'MAX_OPTION_WORKERS_PER_SYMBOL', 8),
        'API_CALLS_PER_SECOND': getattr(yield_module, 'API_CALLS_PER_SECOND', 8),
        'REAL_TIME_UPDATE_INTERVAL': getattr(yield_module, 'REAL_TIME_UPDATE_INTERVAL', 60)
    }
    
    return render_template("settings.html", 
                         title="Settings",
                         config=config_data)